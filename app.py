import json
import math
import uuid
import os
import string
import secrets
import smtplib
from email.message import EmailMessage
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import db  # camada de persistência SQL (Postgres em prod, SQLite local)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'rumo-obras-secret-2024')

DATA_FILE             = 'data/registros.json'
CONTRATOS_CONFIG_FILE = 'data/contratos_config.json'
USUARIOS_FILE         = 'data/usuarios.json'
AUDITORIA_FILE        = 'data/auditoria.json'
MOBILIZACAO_FILE      = 'data/mobilizacao.json'
EXPORT_DIR            = 'exports'
ADMIN_PASSWORD        = os.environ.get('ADMIN_PASSWORD', 'Pipoc@2407')
TIPO_MAO_OBRA_FILE    = os.environ.get('TIPO_MAO_OBRA_FILE',
                        r'C:\Users\Admin\Desktop\PBX\BASES DASHs\CONTRATADAS\TIPO DE MAO DE OBRA.xlsx')

# Garante que os diretórios de dados/exportação existam (deploy limpo no Railway)
os.makedirs('data', exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

# Cria as tabelas SQL e faz seed a partir dos JSONs (idempotente).
# A persistência passa a ser o banco; os JSONs viram o snapshot/seed inicial.
db.init_db()

# ── E-mail (SMTP) — configurável por variáveis de ambiente ──
SMTP_HOST     = os.environ.get('SMTP_HOST', '')
SMTP_PORT     = int(os.environ.get('SMTP_PORT', '587') or 587)
SMTP_USER     = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM     = os.environ.get('SMTP_FROM', SMTP_USER or 'no-reply@tms.local')
SMTP_TLS      = os.environ.get('SMTP_TLS', 'true').lower() != 'false'
RESET_TOKEN_TTL_H = 2  # validade do link de redefinição (horas)


def load_tipo_mao_obra():
    mapping = {}
    try:
        wb = load_workbook(TIPO_MAO_OBRA_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                cargo = str(row[0]).strip().lower()
                tipo_raw = str(row[1]).strip().lower()
                mapping[cargo] = 'direto' if 'diret' in tipo_raw else 'indireto'
    except Exception:
        pass
    return mapping


def classify_tipo(funcao):
    if not funcao:
        return 'classificar'
    return TIPO_MAO_OBRA.get(funcao.strip().lower(), 'classificar')


def load_data():
    return db.load_registros()


def save_data(data):
    db.save_registros(data)


def load_contratos_config():
    return db.load_contratos()


def save_contratos_config(cfg):
    db.save_contratos(cfg)


# ── Financeiro: centros de custo, lançamentos e orçamento ────────────────────
def load_centros_custo():
    return db.load_centros_custo()


def save_centros_custo(items):
    db.save_centros_custo(items)


def load_lancamentos():
    return db.load_lancamentos()


def save_lancamentos(items):
    db.save_lancamentos(items)


def load_orcamentos():
    return db.load_orcamentos()


def save_orcamentos(items):
    db.save_orcamentos(items)


def parse_brl(raw):
    """'1.234.567,89' → 1234567.89 (aceita vazio)."""
    raw = (raw or '').strip().replace('.', '').replace(',', '.')
    try:
        return float(raw) if raw else 0.0
    except ValueError:
        return 0.0


def lancamento_status(l):
    """Status derivado das datas realizadas: pago > faturado > medido."""
    if l.get('data_pgto_real'):
        return 'pago'
    if l.get('data_fat_real'):
        return 'faturado'
    return 'medido'


app.jinja_env.globals['lancamento_status'] = lancamento_status


# ── Usuários ─────────────────────────────────────────────────────────────────
def load_usuarios():
    return db.load_usuarios()


def save_usuarios(usuarios):
    db.save_usuarios(usuarios)


def find_user_by_email(email):
    if not email:
        return None
    email = email.strip().lower()
    return next((u for u in load_usuarios() if u.get('email', '').lower() == email), None)


def find_user_by_id(uid):
    if not uid:
        return None
    return next((u for u in load_usuarios() if u.get('id') == uid), None)


# ── Auditoria (log append-only de inclusão/edição/exclusão) ──────────────────
def load_auditoria():
    return db.load_auditoria()


def audit_log(acao, alvo, detalhe=''):
    """Registra uma ação com data/hora e usuário responsável."""
    try:
        log = load_auditoria()
        log.append({
            'data_hora': datetime.now().isoformat(),
            'usuario':   current_user_label(),
            'acao':      acao,       # ex.: 'criar_registro', 'editar_registro', 'excluir_registro'
            'alvo':      alvo,       # ex.: id ou chave afetada
            'detalhe':   detalhe,
        })
        db.save_auditoria(log)
    except Exception as e:
        app.logger.warning(f'Falha ao gravar auditoria: {e}')


def _audit_descricao(a):
    """Texto curto e legível resumindo a ação registrada na auditoria."""
    acao    = a.get('acao', '')
    detalhe = (a.get('detalhe') or '').strip()
    alvo    = (a.get('alvo') or '').strip()
    _, _, tipo = acao.partition('_')   # criar_registro -> tipo='registro'

    if tipo == 'registro':
        base  = detalhe.split(' — por ')[0]
        parts = [p.strip() for p in base.split('/') if p.strip()]
        if len(parts) >= 3:
            return f'Registro de {parts[0]} — contrato {parts[1]}, semana {parts[2]}'
        return f'Registro ({base})' if base else 'Registro'
    if tipo == 'usuario':
        if acao == 'editar_usuario':
            return f'Permissões do usuário {alvo}'
        return f'Usuário {alvo}'
    if tipo == 'contrato':
        parts = [p.strip() for p in detalhe.split('/') if p.strip()]
        if len(parts) >= 2:
            return f'Contrato {parts[1]} de {parts[0]}'
        return f'Contrato {alvo}'
    return detalhe or alvo or '—'


# ── Sessão / controle de acesso ──────────────────────────────────────────────
def current_user():
    """Retorna o usuário atuante: admin (senha mestra), usuário cadastrado, ou None."""
    if session.get('admin_ok') is True:
        return {'id': 'admin', 'email': 'admin', 'nome': 'Administrador',
                'role': 'admin', 'contratada': None}
    uid = session.get('user_id')
    if uid:
        u = find_user_by_id(uid)
        if u and u.get('ativo', True):
            return u
    return None


def current_user_label():
    u = current_user()
    if not u:
        return 'Público'
    return u.get('nome') or u.get('email') or 'Usuário'


def is_admin():
    return session.get('admin_ok') is True


# Papéis vinculados a uma contratada/contrato (dados filtrados ao próprio contrato):
#   'contratada'    → somente leitura
#   'contratada_rw' → leitura + criação de novos registros (não edita/exclui histórico)
SCOPED_ROLES = ('contratada', 'contratada_rw')

# Papéis com acesso total (admin): master e rumo. 'rumo' = master, exceto que não pode
# alterar usuários master (ver can_manage_user).
ADMIN_ROLES = ('master', 'rumo')

# Coleções apagáveis na tela "Limpar Dados" (rótulo exibido na UI). Fonte única —
# usada tanto em /admin/dados quanto em /admin/limpar-dados.
WIPE_COLS = [
    ('registros', 'Registros semanais'), ('contratos', 'Contratos'),
    ('suprimentos', 'Suprimentos'), ('pacotes', 'Pacotes'),
    ('centros_custo', 'Centros de custo'), ('lancamentos', 'Lançamentos financeiros'),
    ('orcamentos', 'Orçamentos'),
    ('usuarios', 'Usuários'), ('auditoria', 'Auditoria'), ('tms', 'Dados TMS'),
]


def can_write():
    """Admin (senha mestra), 'master', 'rumo' e 'staff' podem lançar/editar/excluir dados."""
    u = current_user()
    return bool(u) and u.get('role') in ('admin', 'master', 'rumo', 'staff')


def can_manage_user(target):
    """Quem pode gerenciar (editar/excluir/etc) o usuário 'target'.
    Admin (senha) e master gerenciam qualquer um. 'rumo' gerencia todos, exceto masters."""
    u = current_user()
    if session.get('admin_ok') is True:
        return True
    if not u or u.get('role') not in ADMIN_ROLES:
        return False
    if u.get('role') == 'rumo' and target and target.get('role') == 'master':
        return False
    return True


def _actor_is_rumo():
    """True se o usuário atuante é 'rumo' (e não admin pela senha mestra)."""
    u = current_user()
    return bool(u) and u.get('role') == 'rumo' and not session.get('admin_ok')


def can_create():
    """Quem pode criar novos registros: quem tem can_write + contratada com lançamento."""
    u = current_user()
    return can_write() or (bool(u) and u.get('role') == 'contratada_rw')


def viewer_contratada():
    """Se o usuário logado é vinculado a uma contratada, retorna o nome dela (filtro). Senão None."""
    u = current_user()
    if u and u.get('role') in SCOPED_ROLES:
        return u.get('contratada')
    return None


def viewer_contrato():
    """Se o usuário logado é vinculado a um contrato específico, retorna-o (filtro). Senão None."""
    u = current_user()
    if u and u.get('role') in SCOPED_ROLES:
        return u.get('contrato')
    return None


def scope_registros(registros):
    """Restringe a lista de registros ao escopo do usuário (contratada + contrato)."""
    vc = viewer_contratada()
    vk = viewer_contrato()
    out = registros
    if vc:
        out = [r for r in out if r.get('contratada') == vc]
    if vk:
        out = [r for r in out if r.get('contrato') == vk]
    return out


# ── E-mail ───────────────────────────────────────────────────────────────────
def _smtp_configured():
    return bool(SMTP_HOST and SMTP_USER and SMTP_PASSWORD)


def send_email(to, subject, html_body):
    """Envia e-mail HTML. Retorna True se enviado; False se SMTP não configurado ou falhou."""
    if not _smtp_configured():
        return False
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = SMTP_FROM
        msg['To'] = to
        msg.set_content('Este e-mail requer um cliente compatível com HTML.')
        msg.add_alternative(html_body, subtype='html')
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
            if SMTP_TLS:
                s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)
        return True
    except Exception as e:
        app.logger.warning(f'Falha ao enviar e-mail para {to}: {e}')
        return False


def _gen_password(n=10):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(n))


def _gen_token():
    return secrets.token_urlsafe(32)


def _set_reset_token(user):
    """Gera token de redefinição com validade e persiste no usuário."""
    token = _gen_token()
    user['reset_token']  = token
    user['reset_expira'] = (datetime.now() + timedelta(hours=RESET_TOKEN_TTL_H)).isoformat()
    return token


def _email_boas_vindas_html(nome, email, senha, login_url, reset_url):
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:auto;color:#1a2b3c;">
      <h2 style="color:#003366;">Bem-vindo(a) ao Monitoramento de Obras — TMS</h2>
      <p>Olá{(' ' + nome) if nome else ''}, sua conta de acesso foi criada.</p>
      <table style="border-collapse:collapse;margin:16px 0;">
        <tr><td style="padding:6px 12px;color:#6b7c93;">E-mail</td>
            <td style="padding:6px 12px;font-weight:bold;">{email}</td></tr>
        <tr><td style="padding:6px 12px;color:#6b7c93;">Senha provisória</td>
            <td style="padding:6px 12px;font-weight:bold;font-family:monospace;font-size:15px;">{senha}</td></tr>
      </table>
      <p>
        <a href="{login_url}" style="background:#00AEEF;color:#fff;text-decoration:none;
           padding:10px 20px;border-radius:8px;font-weight:bold;display:inline-block;">Entrar agora</a>
      </p>
      <p style="font-size:13px;color:#6b7c93;">Recomendamos trocar sua senha no primeiro acesso.
         Use o link abaixo (válido por {RESET_TOKEN_TTL_H}h) para definir uma nova senha:</p>
      <p style="font-size:13px;"><a href="{reset_url}">{reset_url}</a></p>
    </div>"""


def _email_reset_html(reset_url):
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:auto;color:#1a2b3c;">
      <h2 style="color:#003366;">Redefinição de senha — TMS</h2>
      <p>Recebemos um pedido para redefinir sua senha. Clique no botão abaixo
         (válido por {RESET_TOKEN_TTL_H} horas):</p>
      <p>
        <a href="{reset_url}" style="background:#00AEEF;color:#fff;text-decoration:none;
           padding:10px 20px;border-radius:8px;font-weight:bold;display:inline-block;">Redefinir senha</a>
      </p>
      <p style="font-size:13px;color:#6b7c93;">Se você não fez este pedido, ignore este e-mail.</p>
      <p style="font-size:13px;"><a href="{reset_url}">{reset_url}</a></p>
    </div>"""


def contrato_key(contratada, contrato):
    return f"{contratada}||{contrato}"


def _iso_week_date(week_str, weekday):
    """Converte 'YYYY-Wnn' para date no dia da semana indicado (1=segunda ... 7=domingo)."""
    year, week = str(week_str).split('-W')
    return datetime.strptime(f'{year}-W{int(week):02d}-{weekday}', '%G-W%V-%u').date()


def get_monday(date_str):
    if not date_str:
        return date_str
    try:
        if '-W' in date_str:
            return _iso_week_date(date_str, 1).strftime('%Y-%m-%d')
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
        monday = d - timedelta(days=d.weekday())
        return monday.strftime('%Y-%m-%d')
    except Exception:
        return date_str


def _week_to_last_day(week_str):
    """Converte 'YYYY-Wnn' para a data do domingo dessa semana no formato DD/MM/YYYY."""
    try:
        if week_str and '-W' in str(week_str):
            return _iso_week_date(week_str, 7).strftime('%d/%m/%Y')
        return week_str or ''
    except Exception:
        return week_str or ''


def _month_of_week(week_date_str):
    """Converte 'YYYY-MM-DD' para 'YYYY-MM'."""
    d = datetime.strptime(week_date_str, '%Y-%m-%d').date()
    return f'{d.year}-{d.month:02d}'


def format_date_br(date_str):
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
        return d.strftime('%d/%m/%Y')
    except Exception:
        return date_str


def format_currency(value):
    try:
        return f"R$ {float(value):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return "R$ 0,00"


app.jinja_env.filters['date_br'] = format_date_br
app.jinja_env.filters['currency'] = format_currency


def format_date_to_week(date_str):
    try:
        s = str(date_str)
        if not s:
            return ''
        if '-W' in s:
            return s
        d = datetime.strptime(s, '%Y-%m-%d')
        return d.strftime('%G-W%V')
    except Exception:
        return ''

app.jinja_env.filters['date_to_week'] = format_date_to_week


def recompute_forecast_acoes(contratada, contrato, overrides=None):
    """Recalcula o forecast das Ações Notáveis de um contrato.

    O desvio acumulado (planejado − real) das semanas ENTREGUES é
    redistribuído, proporcionalmente ao planejado, SOMENTE entre as semanas
    futuras (após a última semana com realizado). O forecast de semanas
    passadas/presentes NUNCA é alterado pelo sistema (= planejado). Atraso
    aumenta as futuras; adiantamento as reduz. `overrides` permite o reajuste
    manual do forecast pelo usuário (aplicado apenas a semanas futuras)."""
    cfg = load_contratos_config()
    key = contrato_key(contratada, contrato)
    entry = cfg.get(key)
    if not entry or not entry.get('linha_base_acoes'):
        return

    # Realizado por ação/semana ISO, a partir dos registros do contrato
    reais = {}
    for r in load_data():
        if r.get('contratada') != contratada or r.get('contrato') != contrato:
            continue
        wk = format_date_to_week(r.get('semana_referencia', ''))
        if not wk:
            continue
        for acao, val in (r.get('acoes_realizadas') or {}).items():
            try:
                reais.setdefault(acao, {})[wk] = float(val)
            except (TypeError, ValueError):
                continue

    for acao in entry['linha_base_acoes']:
        nome = acao.get('acao', '')
        planejado = {k: float(v or 0) for k, v in (acao.get('semanas') or {}).items()}
        real = reais.get(nome, {})

        # Reajustes manuais TRAVADOS — persistem e nunca são recalculados.
        # Novos ajustes vindos do formulário são mesclados aos já existentes.
        manual = dict(acao.get('forecast_manual') or {})
        if overrides and nome in overrides:
            for w, val in (overrides[nome] or {}).items():
                if w in planejado:
                    try:
                        manual[w] = math.ceil(float(val))
                    except (TypeError, ValueError):
                        pass
        acao['forecast_manual'] = manual

        # fronteira passado/futuro = última semana COM realizado (a que notificou o desvio)
        entregues = [w for w in real if w in planejado]
        limite = max(entregues) if entregues else None
        # distribui SOMENTE o desvio dessa última semana (não o acumulado)
        net = (planejado[limite] - real[limite]) if limite is not None else 0.0
        futuras = [w for w in planejado if limite is not None and w > limite]
        # semanas futuras travadas manualmente (não entram no recálculo)
        manual_fut = {w: float(manual[w]) for w in manual if w in futuras}
        absorvido = sum(manual_fut[w] - planejado[w] for w in manual_fut)
        restante = net - absorvido            # desvio a distribuir entre as NÃO travadas
        auto_fut = [w for w in futuras if w not in manual_fut]
        base = sum(planejado[w] for w in auto_fut)

        forecast = {}
        for w in planejado:
            if w in manual_fut:
                forecast[w] = math.ceil(manual_fut[w])           # travado (manual)
            elif w in auto_fut:
                if base > 0:
                    share = restante * (planejado[w] / base)
                else:
                    share = restante / len(auto_fut) if auto_fut else 0.0
                forecast[w] = math.ceil(max(0.0, planejado[w] + share))
            else:
                forecast[w] = math.ceil(planejado[w])            # passado/presente = planejado
        acao['forecast'] = forecast

    save_contratos_config(cfg)


def acoes_acumulado_map(exclude_id=None):
    """Realizado acumulado por contrato/ação somando todos os registros
    (opcionalmente excluindo um registro — usado na edição para não contar
    duas vezes a semana que está sendo editada). Alimenta a coluna "Evol."
    do formulário: acumulado ÷ total planejado da ação."""
    acum = {}
    for r in load_data():
        if exclude_id and r.get('id') == exclude_id:
            continue
        k = contrato_key(r.get('contratada', ''), r.get('contrato', ''))
        for acao, val in (r.get('acoes_realizadas') or {}).items():
            try:
                acum.setdefault(k, {})
                acum[k][acao] = acum[k].get(acao, 0) + float(val)
            except (TypeError, ValueError):
                continue
    return acum


def acoes_real_semanas_map(exclude_id=None):
    """Realizado por contrato/ação/semana ISO (para a prévia de reajuste do
    forecast no formulário). Exclui opcionalmente um registro (edição)."""
    m = {}
    for r in load_data():
        if exclude_id and r.get('id') == exclude_id:
            continue
        k = contrato_key(r.get('contratada', ''), r.get('contrato', ''))
        wk = format_date_to_week(r.get('semana_referencia', ''))
        if not wk:
            continue
        for acao, val in (r.get('acoes_realizadas') or {}).items():
            try:
                m.setdefault(k, {}).setdefault(acao, {})[wk] = float(val)
            except (TypeError, ValueError):
                continue
    return m


def acoes_forecast_manual_map():
    """Reajustes manuais de forecast travados, por contrato/ação/semana.
    Usado no formulário para exibir as semanas já travadas pelo usuário."""
    m = {}
    for key, entry in load_contratos_config().items():
        for a in entry.get('linha_base_acoes', []):
            fm = a.get('forecast_manual')
            if fm:
                m.setdefault(key, {})[a.get('acao', '')] = fm
    return m


def avanco_fisico_fc_manual_map(exclude_id=None):
    """Forecast manual do Avanço Físico por contratada||contrato → mês → %.
    Agrega o campo 'avanco_fisico_forecast' dos registros mais recentes."""
    result = {}
    for r in load_data():
        if exclude_id and r.get('id') == exclude_id:
            continue
        fc = r.get('avanco_fisico_forecast', {})
        if not fc:
            continue
        key = f"{r.get('contratada', '')}||{r.get('contrato', '')}"
        result.setdefault(key, {}).update(fc)
    return result


def contract_status(data):
    manual = data.get('status_manual', 'auto')
    if manual in ('ativo', 'encerrado'):
        return manual
    fim = data.get('data_fim_contrato', '')
    if not fim:
        return 'ativo'
    try:
        if '-W' in str(fim):
            d = _iso_week_date(fim, 7)
        else:
            d = datetime.strptime(str(fim), '%Y-%m-%d').date()
        return 'encerrado' if date.today() > d else 'ativo'
    except Exception:
        return 'ativo'

app.jinja_env.globals['contract_status'] = contract_status


def valor_aditivos(cdata):
    """Soma dos aditivos de VALOR do contrato."""
    return sum(float(a.get('valor', 0) or 0)
               for a in (cdata.get('aditivos') or []) if a.get('tipo') == 'valor')


def valor_efetivo(cdata):
    """Valor contratado efetivo = valor base + aditivos de valor.
    Usado em todos os KPIs financeiros (Financeiro, Dashboard, Consolidado)."""
    return float(cdata.get('valor_contrato', 0) or 0) + valor_aditivos(cdata)


app.jinja_env.globals['valor_efetivo'] = valor_efetivo
app.jinja_env.globals['valor_aditivos'] = valor_aditivos


_ALL_ROLES = [
    ('master',        'Master',              'rgba(155,89,182,.8)'),
    ('rumo',          'RUMO',                'rgba(0,174,239,.8)'),
    ('staff',         'Equipe interna',      'rgba(0,174,239,.6)'),
    ('contratada_rw', 'Contratada · RW',     'rgba(70,194,106,.8)'),
    ('contratada',    'Contratada · leitura','rgba(70,194,106,.6)'),
    ('mobilizacao',   'Mobilização',         'rgba(255,112,67,.8)'),
]


@app.context_processor
def inject_now():
    u = current_user()
    preview = session.get('preview_role') if u and u.get('role') in ('admin', 'master', 'rumo') else None
    view_role = preview or (u.get('role') if u else None)
    return {
        'now': datetime.now(),
        'current_user': u,
        'can_write': can_write(),
        'can_create': can_create(),
        'viewer_contratada': viewer_contratada(),
        'viewer_contrato': viewer_contrato(),
        'view_role': view_role,
        'preview_role': preview,
        'all_roles': _ALL_ROLES,
    }


@app.route('/preview-role/<role>')
def set_preview_role(role):
    u = current_user()
    if u and u.get('role') in ('admin', 'master', 'rumo'):
        valid = {r[0] for r in _ALL_ROLES}
        if role in valid:
            session['preview_role'] = role
    return redirect(request.referrer or url_for('capa'))


@app.route('/preview-role/reset')
def reset_preview_role():
    session.pop('preview_role', None)
    return redirect(request.referrer or url_for('capa'))


# Endpoints acessíveis sem login (capa, autenticação e assets)
PUBLIC_ENDPOINTS = {
    'capa', 'login', 'logout', 'esqueci_senha', 'redefinir_senha',
    'admin_login', 'static',
}


_MOB_ENDPOINTS = {'mobilizacao', 'mobilizacao_config', 'mobilizacao_documentos', 'static', 'logout', 'capa',
                  'mobilizacao_exportar_excel', 'mobilizacao_importar_excel'}


@app.before_request
def _require_login_global():
    """Bloqueia o acesso a qualquer página (exceto a capa e fluxo de login) sem estar logado."""
    endpoint = request.endpoint
    if endpoint is None or endpoint in PUBLIC_ENDPOINTS:
        return
    u = current_user()
    if u is None:
        if endpoint.startswith('admin'):
            return redirect(url_for('admin_login'))
        flash('Faça login para acessar o sistema.', 'warning')
        return redirect(url_for('login', next=request.path))
    # Perfil MOBILIZAÇÃO: acesso restrito às páginas do módulo
    if u.get('role') == 'mobilizacao' and endpoint not in _MOB_ENDPOINTS:
        return redirect(url_for('mobilizacao'))


def get_contratadas():
    """Retorna lista ordenada de contratadas cadastradas no Admin (contratos_config.json)."""
    cfg = load_contratos_config()
    return sorted({v.get('contratada', '') for v in cfg.values() if v.get('contratada')})


_FUNCOES_PADRAO = [
    # ── Gestão & Técnico ──
    {'cargo': 'Engenheiro Civil',           'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Engenheiro de Segurança',    'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Técnico de Segurança',       'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Mestre de Obras',            'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Encarregado',                'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Apontador',                  'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Almoxarife',                 'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    {'cargo': 'Topógrafo',                  'tipo': 'indireto', 'grupo': 'Gestão & Técnico'},
    # ── Mão de Obra Direta ──
    {'cargo': 'Pedreiro',                   'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Armador',                    'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Carpinteiro',                'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Eletricista',                'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Soldador',                   'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Serralheiro',                'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Pintor',                     'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Servente',                   'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Ajudante Geral',             'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    {'cargo': 'Sinaleiro',                  'tipo': 'direto', 'grupo': 'Mão de Obra Direta'},
    # ── Operadores ──
    {'cargo': 'Operador de Escavadeira',    'tipo': 'direto', 'grupo': 'Operadores'},
    {'cargo': 'Operador de Máquinas',       'tipo': 'direto', 'grupo': 'Operadores'},
    {'cargo': 'Operador de Pavimentadora',  'tipo': 'direto', 'grupo': 'Operadores'},
    {'cargo': 'Operador de Compactador',    'tipo': 'direto', 'grupo': 'Operadores'},
    {'cargo': 'Operador de Guindaste',      'tipo': 'direto', 'grupo': 'Operadores'},
    {'cargo': 'Motorista',                  'tipo': 'direto', 'grupo': 'Operadores'},
    # ── Maquinário ──
    {'cargo': 'Escavadeira Hidráulica',     'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Retroescavadeira',           'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Pá Carregadeira',            'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Motoniveladora (Patrol)',     'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Trator de Esteiras',         'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Pavimentadora Asfáltica',    'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Rolo Compactador Liso',      'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Rolo Compactador Pé de Carneiro', 'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Caminhão Basculante',        'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Caminhão Betoneira',         'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Caminhão Munck',             'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Caminhão Pipa',              'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Guindaste Telescópico',      'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Betoneira',                  'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Bomba de Concreto',          'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Compressor de Ar',           'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Gerador de Energia',         'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Plataforma Elevatória',      'tipo': 'equipamento', 'grupo': 'Maquinário'},
    {'cargo': 'Minicarregadeira (Bob Cat)', 'tipo': 'equipamento', 'grupo': 'Maquinário'},
]

def get_funcoes_list():
    """Carrega lista de cargos/tipos do Excel; usa lista padrão se arquivo não encontrado."""
    result = []
    try:
        wb = load_workbook(TIPO_MAO_OBRA_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cargo    = str(row[0]).strip()
                tipo_raw = str(row[1]).strip().lower() if row[1] else ''
                tipo     = 'direto' if 'diret' in tipo_raw else 'indireto'
                result.append({'cargo': cargo, 'tipo': tipo})
    except Exception:
        pass
    return result if result else _FUNCOES_PADRAO


# Mapa cargo→tipo para classificação. Usa o Excel se existir; senão deriva da lista padrão.
TIPO_MAO_OBRA = load_tipo_mao_obra()
if not TIPO_MAO_OBRA:
    TIPO_MAO_OBRA = {
        f['cargo'].strip().lower(): f['tipo']
        for f in _FUNCOES_PADRAO if f['tipo'] in ('direto', 'indireto')
    }

PLUVIOMETRIA_OPCOES = [
    "Tempo Bom",
    "Chuva Produtiva",
    "Chuva Parcial",
    "Chuva Improdutiva",
    "Incidência de Raio",
    "Sem Expediente",
]

# Estimativa de precipitação diária (mm) por estado qualitativo — usada para
# converter os registros de pluviometria (qualitativos) em uma série semanal (mm)
# no gráfico "Pluviometria Semanal" do dashboard.
PLUVIOMETRIA_MM = {
    "Tempo Bom":          0,
    "Chuva Produtiva":    4,
    "Chuva Parcial":      12,
    "Chuva Improdutiva":  28,
    "Incidência de Raio": 20,
    "Sem Expediente":     0,
}

DIAS_SEMANA = [
    ('segunda', 'Segunda'),
    ('terca',   'Terça'),
    ('quarta',  'Quarta'),
    ('quinta',  'Quinta'),
    ('sexta',   'Sexta'),
    ('sabado',  'Sábado'),
    ('domingo', 'Domingo'),
]

def parse_efetivo(form):
    funcoes = form.getlist('efetivo_funcao[]')
    quantidades = form.getlist('efetivo_quantidade[]')
    efetivo = []
    total_direto = 0
    total_indireto = 0
    for funcao, qtd_str in zip(funcoes, quantidades):
        if funcao.strip():
            try:
                qtd = int(qtd_str)
            except Exception:
                qtd = 0
            tipo = classify_tipo(funcao)
            efetivo.append({'funcao': funcao.strip(), 'quantidade': qtd, 'tipo': tipo})
            if tipo == 'direto':
                total_direto += qtd
            elif tipo == 'indireto':
                total_indireto += qtd
    return efetivo, total_direto, total_indireto


def parse_equipamentos(form):
    descricoes = form.getlist('equip_descricao[]')
    quantidades = form.getlist('equip_quantidade[]')
    equipamentos = []
    for desc, qtd_str in zip(descricoes, quantidades):
        if desc.strip():
            try:
                qtd = int(qtd_str)
            except Exception:
                qtd = 0
            equipamentos.append({'descricao': desc.strip(), 'quantidade': qtd})
    return equipamentos


def parse_pluviometria(form):
    return {dia: form.get(f'pluv_{dia}', '').strip() for dia, _ in DIAS_SEMANA}


def _parse_json_dict(form, field):
    try:
        data = json.loads(form.get(field, '{}'))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def parse_acoes_realizadas(form):
    return _parse_json_dict(form, 'acoes_realizadas_json')


def fin_base_curve(cfg, filtro_contratada, semanas_sorted, filtro_contrato=None, replan=False):
    """Linha de base financeira acumulada (mensal) alinhada ao eixo semanal do gráfico.

    Se replan=True, usa por contrato o replanejamento mais recente sobreposto ao
    planejado original e retorna (curva, houve_replan)."""
    monthly = {}
    houve = False
    for _, cdata in cfg.items():
        if filtro_contratada and cdata.get('contratada') != filtro_contratada:
            continue
        if filtro_contrato and cdata.get('contrato') != filtro_contrato:
            continue
        base = {entry['semana']: float(entry.get('valor', 0) or 0)
                for entry in cdata.get('linha_base_financeira', []) if entry.get('semana')}
        if replan:
            base, ch = _mesclar_replan(base, cdata.get('lb_fin_replanejados'))
            houve = houve or ch
        for m, v in base.items():
            monthly[m] = monthly.get(m, 0) + float(v or 0)

    cumul, acum = {}, 0
    for m in sorted(monthly):
        acum += monthly[m]
        cumul[m] = round(acum, 2)
    months = sorted(cumul)

    curve = []
    for s in semanas_sorted:
        m = _month_of_week(s)
        curve.append(next((cumul[bm] for bm in reversed(months) if bm <= m), None))
    return (curve, houve) if replan else curve


# ── Helpers das Curvas S (Forecast + Replanejamento) ─────────────────────────
def _iter_months(start_ym, end_ym):
    """Lista 'YYYY-MM' de start_ym a end_ym (inclusive)."""
    sy, sm = (int(x) for x in start_ym.split('-'))
    ey, em = (int(x) for x in end_ym.split('-'))
    cur, end = sy * 12 + (sm - 1), ey * 12 + (em - 1)
    out = []
    while cur <= end:
        out.append(f'{cur // 12:04d}-{cur % 12 + 1:02d}')
        cur += 1
    return out


def _future_dates(cfg, semanas_sorted, filtro_contratada=None, filtro_contrato=None):
    """Datas mensais futuras (YYYY-MM-28) após a última semana com dados, até o fim
    do baseline/contrato — estende o eixo do tempo para o forecast/replan projetarem."""
    meses = set()
    for c in cfg.values():
        if filtro_contratada and c.get('contratada') != filtro_contratada:
            continue
        if filtro_contrato and c.get('contrato') != filtro_contrato:
            continue
        for it in c.get('linha_base_financeira', []):
            if it.get('semana'):
                meses.add(it['semana'])
        for it in c.get('linha_base_fisica', []):
            if it.get('semana'):
                meses.add(it['semana'])
        _fim = c.get('data_fim_contrato', '')
        if _fim:
            try:
                meses.add(_month_of_week(_iso_week_date(_fim, 1).strftime('%Y-%m-%d')))
            except Exception:
                pass
    out = []
    if semanas_sorted and meses:
        end_month = max(meses)
        last_hist_month = _month_of_week(semanas_sorted[-1])
        if end_month > last_hist_month:
            for m in _iter_months(last_hist_month, end_month):
                if m > last_hist_month:
                    out.append(f'{m}-28')
    return out


def _forecast_series(real, base, clamp_max=None):
    """Projeção = linha de base deslocada pelo desvio atual (realizado − previsto na
    última semana com dados). Atrasado → corre abaixo da base; adiantado → acima."""
    n = len(base)
    fc = [None] * n
    li = next((i for i in range(len(real) - 1, -1, -1) if real[i] is not None), None)
    if li is None:
        return fc
    gap = (real[li] - base[li]) if base[li] is not None else 0.0
    for i in range(li, n):
        if base[i] is None:
            continue
        v = max(base[i] + gap, 0)
        if clamp_max is not None:
            v = min(v, clamp_max)
        fc[i] = round(v, 2)
    return fc


def _mesclar_replan(base_dict, replans):
    """Sobrepõe o replanejamento mais recente ao planejado original.
    Retorna (dict_revisado, houve_replan)."""
    revisado = {k: float(v or 0) for k, v in base_dict.items()}
    repls = [r for r in (replans or []) if r]
    if not repls:
        return revisado, False
    for k, v in repls[-1].items():
        revisado[k] = float(v or 0)
    return revisado, True


def _parse_baseline_mensal(form, data):
    """Lê os campos das Linhas de Base mensais (Financeira + Física) do formulário e
    grava em `data`. Mesma lógica usada em Configurar Contrato e Configurações TMS."""
    # ── Financeira ──
    try:
        lb_fin_plan = json.loads(form.get('lb_fin_json', '{}') or '{}')
    except (json.JSONDecodeError, ValueError):
        lb_fin_plan = {}
    data['lb_fin_planejado'] = lb_fin_plan
    # Mantém formato legado para compatibilidade com dashboards
    data['linha_base_financeira'] = [
        {'semana': m, 'valor': float(v)}
        for m, v in sorted(lb_fin_plan.items()) if v is not None
    ]
    try:
        data['lb_fin_forecast'] = json.loads(form.get('lb_fin_fcst_json', '{}') or '{}')
    except (json.JSONDecodeError, ValueError):
        data['lb_fin_forecast'] = {}
    try:
        data['lb_fin_replanejados'] = json.loads(form.get('lb_fin_repls_json', '[]') or '[]')
    except (json.JSONDecodeError, ValueError):
        data['lb_fin_replanejados'] = []
    try:
        data['lb_fin_repl_n'] = max(0, min(3, int(form.get('lb_fin_repl_n', '0') or 0)))
    except (TypeError, ValueError):
        data['lb_fin_repl_n'] = 0
    try:
        data['lb_fin_extra'] = json.loads(form.get('lb_fin_extra_json', '[]') or '[]')
    except (json.JSONDecodeError, ValueError):
        data['lb_fin_extra'] = []

    # ── Física ──
    try:
        lb_fis_plan = json.loads(form.get('lb_fis_json', '{}') or '{}')
    except (json.JSONDecodeError, ValueError):
        lb_fis_plan = {}
    data['lb_fis_planejado'] = lb_fis_plan
    data['linha_base_fisica'] = [
        {'semana': m, 'percentual': float(v)}
        for m, v in sorted(lb_fis_plan.items()) if v is not None
    ]
    try:
        data['lb_fis_forecast'] = json.loads(form.get('lb_fis_fcst_json', '{}') or '{}')
    except (json.JSONDecodeError, ValueError):
        data['lb_fis_forecast'] = {}
    try:
        data['lb_fis_replanejados'] = json.loads(form.get('lb_fis_repls_json', '[]') or '[]')
    except (json.JSONDecodeError, ValueError):
        data['lb_fis_replanejados'] = []
    try:
        data['lb_fis_repl_n'] = max(0, min(3, int(form.get('lb_fis_repl_n', '0') or 0)))
    except (TypeError, ValueError):
        data['lb_fis_repl_n'] = 0
    try:
        data['lb_fis_extra'] = json.loads(form.get('lb_fis_extra_json', '[]') or '[]')
    except (json.JSONDecodeError, ValueError):
        data['lb_fis_extra'] = []


@app.route('/')
def capa():
    return render_template('capa.html')


@app.route('/financeiro')
def financeiro():
    registros = load_data()
    todas_contratadas = sorted(set(r.get('contratada', '') for r in registros if r.get('contratada')))

    filtro_contratada = request.args.get('contratada', '')
    filtro_de         = request.args.get('de', '')
    filtro_ate        = request.args.get('ate', '')

    vc = viewer_contratada()
    if vc:
        filtro_contratada = vc
        todas_contratadas = [vc]
    vk = viewer_contrato()   # contrato do usuário (segmentação)

    reg = scope_registros(registros[:])
    if filtro_contratada:
        reg = [r for r in reg if r.get('contratada') == filtro_contratada]
    if filtro_de:
        reg = [r for r in reg if r.get('semana_referencia', '') >= filtro_de]
    if filtro_ate:
        reg = [r for r in reg if r.get('semana_referencia', '') <= filtro_ate]

    cfg = load_contratos_config()

    # ── Dados financeiros complementares (lançamentos, centros de custo, orçamento) ──
    centros  = load_centros_custo()
    cc_by_id = {c.get('id'): c for c in centros}
    orcs     = load_orcamentos()
    lanc_scope = [l for l in load_lancamentos()
                  if (not filtro_contratada or l.get('contratada') == filtro_contratada)
                  and (not vk or l.get('contrato') == vk)]

    _vazio = dict(kpis=None, contratadas=todas_contratadas,
                  filtro_contratada=filtro_contratada,
                  filtro_de=filtro_de, filtro_ate=filtro_ate,
                  chart_labels='[]', curva_fin_acum='[]', curva_fin_base='[]',
                  contratos_fin=[], bar_labels='[]', bar_vals='[]', bar_colors='[]',
                  evm=None, funil='[]', funil_labels='[]',
                  cf_labels='[]', cf_pago='[]', cf_prev='[]', cf_proj='[]', cf_acum='[]',
                  class_data='{}', cc_labels='[]', cc_orcado='[]',
                  cc_comprometido='[]', cc_medido='[]',
                  n_lancamentos=len(lanc_scope), tem_centros=bool(centros))

    if not reg:
        return render_template('financeiro.html', **_vazio)

    reg_ord = sorted(reg, key=lambda r: r.get('semana_referencia', ''))

    # ── KPIs globais (valor total considera todos os contratos configurados) ──
    total_valor_contrato = sum(
        valor_efetivo(cdata)
        for cdata in cfg.values()
        if (not filtro_contratada or cdata.get('contratada') == filtro_contratada)
        and (not vk or cdata.get('contrato') == vk)
    )
    total_medido  = sum(r.get('valor_medido', 0) for r in reg)
    saldo_global  = total_valor_contrato - total_medido
    pct_global    = round(total_medido / total_valor_contrato * 100, 1) if total_valor_contrato else 0

    ultima_semana = reg_ord[-1].get('semana_referencia', '')
    reg_semana    = [r for r in reg if r.get('semana_referencia') == ultima_semana]
    valor_semana  = sum(r.get('valor_medido', 0) for r in reg_semana)

    # ── Dados semanais para Curva S ──
    semanas_data = {}
    for r in reg_ord:
        sem = r.get('semana_referencia', '')
        semanas_data[sem] = semanas_data.get(sem, 0) + r.get('valor_medido', 0)

    semanas_sorted = sorted(semanas_data.keys())
    chart_labels   = [format_date_br(s) for s in semanas_sorted]

    curva_fin_acum, acum = [], 0
    for s in semanas_sorted:
        acum += semanas_data[s]
        curva_fin_acum.append(round(acum, 2))

    curva_fin_base = fin_base_curve(cfg, filtro_contratada, semanas_sorted, vk)

    # ── Pagamentos/faturamentos por contrato (dos lançamentos) ──
    pago_por_par, fat_por_par = {}, {}
    for l in lanc_scope:
        par = (l.get('contratada', ''), l.get('contrato', ''))
        vl  = float(l.get('valor_liquido') or 0)
        if l.get('data_pgto_real'):
            pago_por_par[par] = pago_por_par.get(par, 0) + vl
        if l.get('data_fat_real'):
            fat_por_par[par] = fat_por_par.get(par, 0) + vl

    # ── Tabela por contrato + agregações CAPEX/OPEX e centro de custo ──
    contratos_fin = []
    class_data = {'CAPEX': {'comprometido': 0, 'medido': 0, 'pago': 0},
                  'OPEX':  {'comprometido': 0, 'medido': 0, 'pago': 0}}
    cc_data = {}   # nome do centro → {orcado, comprometido, medido}
    for _, cdata in sorted(cfg.items()):
        contratada = cdata.get('contratada', '')
        contrato   = cdata.get('contrato', '')
        if filtro_contratada and contratada != filtro_contratada:
            continue
        if vk and contrato != vk:
            continue
        valor  = valor_efetivo(cdata)
        reg_c  = [r for r in reg if r.get('contratada') == contratada and r.get('contrato') == contrato]
        medido = sum(r.get('valor_medido', 0) for r in reg_c)
        saldo  = valor - medido
        pct    = round(medido / valor * 100, 1) if valor else 0
        classe = 'OPEX' if (cdata.get('classificacao') or 'CAPEX').upper() == 'OPEX' else 'CAPEX'
        cc     = cc_by_id.get(cdata.get('centro_custo') or '', {})
        pago_c = pago_por_par.get((contratada, contrato), 0)
        contratos_fin.append({
            'contratada': contratada,
            'contrato':   contrato,
            'valor':      valor,
            'valor_base': float(cdata.get('valor_contrato', 0) or 0),
            'aditivos':   valor_aditivos(cdata),
            'medido':     medido,
            'saldo':      saldo,
            'pct':        pct,
            'pago':       pago_c,
            'classificacao': classe,
            'centro_custo':  cc.get('nome', ''),
            'status':     contract_status(cdata),
        })
        class_data[classe]['comprometido'] += valor
        class_data[classe]['medido']       += medido
        class_data[classe]['pago']         += pago_c
        cc_nome = cc.get('nome') or 'Sem centro de custo'
        cc_data.setdefault(cc_nome, {'orcado': 0, 'comprometido': 0, 'medido': 0})
        cc_data[cc_nome]['comprometido'] += valor
        cc_data[cc_nome]['medido']       += medido

    # Orçamento por centro de custo (grade mensal somada)
    orcado_total = 0
    for o in orcs:
        cc_nome = cc_by_id.get(o.get('centro_custo') or '', {}).get('nome') or 'Sem centro de custo'
        v = float(o.get('valor') or 0)
        if not filtro_contratada or cc_nome in cc_data:
            orcado_total += v
        if cc_nome in cc_data:
            cc_data[cc_nome]['orcado'] += v
        elif not filtro_contratada:
            cc_data.setdefault(cc_nome, {'orcado': 0, 'comprometido': 0, 'medido': 0})
            cc_data[cc_nome]['orcado'] += v

    # ── EVM (PV / EV / AC → CPI, SPI, EAC) ──
    hoje_m = date.today().strftime('%Y-%m')
    pv = 0
    for cdata in cfg.values():
        if filtro_contratada and cdata.get('contratada') != filtro_contratada:
            continue
        if vk and cdata.get('contrato') != vk:
            continue
        for e in cdata.get('linha_base_financeira', []):
            if e.get('semana') and e['semana'] <= hoje_m:
                pv += float(e.get('valor', 0) or 0)
    ev = 0
    _af_ultimo = {}   # (contratada, contrato) → último avanço físico
    for r in reg_ord:
        if r.get('avanco_fisico') is not None:
            _af_ultimo[(r.get('contratada', ''), r.get('contrato', ''))] = float(r.get('avanco_fisico') or 0)
    for cdata in cfg.values():
        par = (cdata.get('contratada', ''), cdata.get('contrato', ''))
        if filtro_contratada and par[0] != filtro_contratada:
            continue
        if vk and par[1] != vk:
            continue
        ev += (_af_ultimo.get(par, 0) / 100.0) * valor_efetivo(cdata)
    ac  = total_medido
    cpi = round(ev / ac, 2) if ac else None
    spi = round(ev / pv, 2) if pv else None
    eac = round(total_valor_contrato / cpi, 2) if cpi else None
    evm = dict(pv=pv, ev=ev, ac=ac, cpi=cpi, spi=spi, eac=eac,
               var_custo=ev - ac, var_prazo=ev - pv)

    # ── Fluxo de caixa mensal ──
    def _shift_month(ym, n):
        y, m = int(ym[:4]), int(ym[5:7])
        t = y * 12 + (m - 1) + n
        return f'{t // 12:04d}-{t % 12 + 1:02d}'

    cf_pago_m, cf_prev_m, cf_proj_m = {}, {}, {}
    for l in lanc_scope:
        vl = float(l.get('valor_liquido') or 0)
        if l.get('data_pgto_real'):
            m = l['data_pgto_real'][:7]
            cf_pago_m[m] = cf_pago_m.get(m, 0) + vl
        elif l.get('data_pgto_prev'):
            m = l['data_pgto_prev'][:7]
            cf_prev_m[m] = cf_prev_m.get(m, 0) + vl
    # Projeção: baseline financeira mensal deslocada pelo prazo de pagamento do contrato
    for cdata in cfg.values():
        if filtro_contratada and cdata.get('contratada') != filtro_contratada:
            continue
        if vk and cdata.get('contrato') != vk:
            continue
        shift = round((cdata.get('prazo_pagamento_dias') or 30) / 30)
        fator = 1 - (float(cdata.get('retencao_pct') or 0) + float(cdata.get('impostos_pct') or 0)) / 100.0
        for e in cdata.get('linha_base_financeira', []):
            if not e.get('semana'):
                continue
            m2 = _shift_month(e['semana'], shift)
            if m2 > hoje_m and m2 not in cf_prev_m:
                cf_proj_m[m2] = cf_proj_m.get(m2, 0) + float(e.get('valor', 0) or 0) * fator
    cf_months = sorted(set(cf_pago_m) | set(cf_prev_m) | set(cf_proj_m))
    cf_labels, cf_pago_l, cf_prev_l, cf_proj_l, cf_acum_l = [], [], [], [], []
    _ac = 0
    for m in cf_months:
        cf_labels.append(f'{m[5:7]}/{m[:4]}')
        p, pr, pj = cf_pago_m.get(m, 0), cf_prev_m.get(m, 0), cf_proj_m.get(m, 0)
        cf_pago_l.append(round(p, 2))
        cf_prev_l.append(round(pr, 2))
        cf_proj_l.append(round(pj, 2))
        _ac += p + pr + pj
        cf_acum_l.append(round(_ac, 2))

    # ── Funil orçado → comprometido → medido → faturado → pago ──
    faturado_total = sum(fat_por_par.values())
    pago_total     = sum(pago_por_par.values())
    retido_acum    = sum(float(l.get('valor_bruto') or 0) * float(l.get('retencao_pct') or 0) / 100.0
                         for l in lanc_scope)
    funil_labels = ['Orçado', 'Comprometido', 'Medido', 'Faturado', 'Pago']
    funil_vals   = [round(orcado_total, 2), round(total_valor_contrato, 2),
                    round(total_medido, 2), round(faturado_total, 2), round(pago_total, 2)]

    cc_sorted = sorted(cc_data.items(), key=lambda x: -x[1]['comprometido'])

    # ── Barra: medição por contratada ──
    med_por_c = {}
    for r in reg:
        c = r.get('contratada', '')
        if c:
            med_por_c[c] = med_por_c.get(c, 0) + r.get('valor_medido', 0)
    med_sorted  = sorted(med_por_c.items(), key=lambda x: -x[1])
    bar_labels  = json.dumps([x[0] for x in med_sorted])
    bar_vals    = json.dumps([round(x[1], 2) for x in med_sorted])
    _PALETTE    = ['rgba(0,212,255,.85)', 'rgba(141,198,63,.85)', 'rgba(240,165,0,.85)',
                   'rgba(255,107,122,.85)', 'rgba(155,89,182,.85)']
    bar_colors  = json.dumps([_PALETTE[i % len(_PALETTE)] for i in range(len(med_sorted))])

    total_base_soma = sum(c['valor_base'] for c in contratos_fin)
    kpis = dict(
        total_valor_contrato=total_valor_contrato,
        total_base=total_base_soma,
        total_aditivos=total_valor_contrato - total_base_soma,
        total_medido=total_medido,
        saldo=saldo_global,
        pct_global=pct_global,
        valor_semana=valor_semana,
        ultima_semana=ultima_semana,
        faturado=faturado_total,
        pago=pago_total,
        retido=retido_acum,
        orcado=orcado_total,
    )

    return render_template('financeiro.html',
                           kpis=kpis,
                           contratadas=todas_contratadas,
                           filtro_contratada=filtro_contratada,
                           filtro_de=filtro_de,
                           filtro_ate=filtro_ate,
                           chart_labels=json.dumps(chart_labels),
                           curva_fin_acum=json.dumps(curva_fin_acum),
                           curva_fin_base=json.dumps(curva_fin_base),
                           contratos_fin=contratos_fin,
                           bar_labels=bar_labels,
                           bar_vals=bar_vals,
                           bar_colors=bar_colors,
                           evm=evm,
                           funil=json.dumps(funil_vals),
                           funil_labels=json.dumps(funil_labels),
                           cf_labels=json.dumps(cf_labels),
                           cf_pago=json.dumps(cf_pago_l),
                           cf_prev=json.dumps(cf_prev_l),
                           cf_proj=json.dumps(cf_proj_l),
                           cf_acum=json.dumps(cf_acum_l),
                           class_data=json.dumps(class_data),
                           cc_labels=json.dumps([x[0] for x in cc_sorted]),
                           cc_orcado=json.dumps([round(x[1]['orcado'], 2) for x in cc_sorted]),
                           cc_comprometido=json.dumps([round(x[1]['comprometido'], 2) for x in cc_sorted]),
                           cc_medido=json.dumps([round(x[1]['medido'], 2) for x in cc_sorted]),
                           n_lancamentos=len(lanc_scope),
                           tem_centros=bool(centros))


# ── Lançamentos financeiros (medição → fatura → pagamento) ───────────────────
@app.route('/financeiro/lancamentos')
def financeiro_lancamentos():
    cfg  = load_contratos_config()
    lanc = load_lancamentos()

    vc, vk = viewer_contratada(), viewer_contrato()
    f_contratada = vc or request.args.get('contratada', '')
    f_contrato   = vk or request.args.get('contrato', '')
    f_status     = request.args.get('status', '')

    itens = [l for l in lanc
             if (not f_contratada or l.get('contratada') == f_contratada)
             and (not f_contrato or l.get('contrato') == f_contrato)]
    if f_status:
        itens = [l for l in itens if lancamento_status(l) == f_status]
    itens.sort(key=lambda l: (l.get('competencia', ''), l.get('criado_em', '')), reverse=True)

    # Contratos disponíveis para o form (escopo do usuário)
    contratos_opts = []
    for k, c in sorted(cfg.items()):
        if vc and c.get('contratada') != vc:
            continue
        if vk and c.get('contrato') != vk:
            continue
        contratos_opts.append({
            'key': k, 'contratada': c.get('contratada', ''), 'contrato': c.get('contrato', ''),
            'retencao_pct': float(c.get('retencao_pct') or 0),
            'impostos_pct': float(c.get('impostos_pct') or 0),
            'prazo': int(c.get('prazo_pagamento_dias') or 30),
        })
    tot_bruto   = sum(float(l.get('valor_bruto') or 0) for l in itens)
    tot_liquido = sum(float(l.get('valor_liquido') or 0) for l in itens)
    contratadas_list = sorted({c.get('contratada', '') for c in cfg.values() if c.get('contratada')})
    if vc:
        contratadas_list = [vc]
    return render_template('financeiro_lancamentos.html',
                           lancamentos=itens, contratos_opts=contratos_opts,
                           contratadas_list=contratadas_list,
                           f_contratada=f_contratada, f_contrato=f_contrato, f_status=f_status,
                           tot_bruto=tot_bruto, tot_liquido=tot_liquido)


def _lancamento_from_form(form):
    bruto = parse_brl(form.get('valor_bruto', ''))
    ret   = parse_brl(form.get('retencao_pct', ''))
    imp   = parse_brl(form.get('impostos_pct', ''))
    key   = form.get('contrato_key', '')
    parts = key.split('||', 1)
    return {
        'contratada':    parts[0] if parts else '',
        'contrato':      parts[1] if len(parts) > 1 else '',
        'competencia':   form.get('competencia', ''),
        'descricao':     form.get('descricao', '').strip(),
        'valor_bruto':   bruto,
        'retencao_pct':  ret,
        'impostos_pct':  imp,
        'valor_liquido': round(bruto * (1 - (ret + imp) / 100.0), 2),
        'data_medicao':  form.get('data_medicao', ''),
        'data_fat_prev': form.get('data_fat_prev', ''),
        'data_fat_real': form.get('data_fat_real', ''),
        'data_pgto_prev': form.get('data_pgto_prev', ''),
        'data_pgto_real': form.get('data_pgto_real', ''),
    }


@app.route('/financeiro/lancamentos/novo', methods=['POST'])
def financeiro_lancamento_novo():
    if not can_write():
        flash('Sem permissão para criar lançamentos.', 'danger')
        return redirect(url_for('financeiro_lancamentos'))
    novo = _lancamento_from_form(request.form)
    if not novo['contratada'] or not novo['competencia'] or not novo['valor_bruto']:
        flash('Preencha contrato, competência e valor bruto.', 'warning')
        return redirect(url_for('financeiro_lancamentos'))
    novo['id'] = str(uuid.uuid4())
    novo['criado_em'] = datetime.now().isoformat()
    novo['criado_por'] = current_user_label()
    lanc = load_lancamentos()
    lanc.append(novo)
    save_lancamentos(lanc)
    audit_log('criar_lancamento', novo['id'],
              f"{novo['contratada']} / {novo['contrato']} — {novo['competencia']} — R$ {novo['valor_bruto']:.2f}")
    flash('Lançamento criado.', 'success')
    return redirect(url_for('financeiro_lancamentos'))


@app.route('/financeiro/lancamentos/<lid>/editar', methods=['POST'])
def financeiro_lancamento_editar(lid):
    if not can_write():
        flash('Sem permissão para editar lançamentos.', 'danger')
        return redirect(url_for('financeiro_lancamentos'))
    lanc = load_lancamentos()
    item = next((l for l in lanc if l.get('id') == lid), None)
    if not item:
        flash('Lançamento não encontrado.', 'danger')
        return redirect(url_for('financeiro_lancamentos'))
    item.update(_lancamento_from_form(request.form))
    item['atualizado_em'] = datetime.now().isoformat()
    item['alterado_por'] = current_user_label()
    save_lancamentos(lanc)
    audit_log('editar_lancamento', lid, f"{item['contratada']} / {item['contrato']} — {item['competencia']}")
    flash('Lançamento atualizado.', 'success')
    return redirect(url_for('financeiro_lancamentos'))


@app.route('/financeiro/lancamentos/<lid>/excluir', methods=['POST'])
def financeiro_lancamento_excluir(lid):
    if not can_write():
        flash('Sem permissão para excluir lançamentos.', 'danger')
        return redirect(url_for('financeiro_lancamentos'))
    lanc = load_lancamentos()
    item = next((l for l in lanc if l.get('id') == lid), None)
    if item:
        lanc.remove(item)
        save_lancamentos(lanc)
        audit_log('excluir_lancamento', lid, f"{item.get('contratada','')} / {item.get('contrato','')}")
        flash('Lançamento excluído.', 'success')
    return redirect(url_for('financeiro_lancamentos'))


# ── Config financeira: centros de custo + orçamento mensal ───────────────────
def _fin_config_required():
    u = current_user()
    return session.get('admin_ok') or (u and u.get('role') in ADMIN_ROLES)


@app.route('/financeiro/config')
def financeiro_config():
    if not _fin_config_required():
        flash('Acesso restrito à configuração financeira.', 'danger')
        return redirect(url_for('financeiro'))
    centros = load_centros_custo()
    orcs    = load_orcamentos()
    try:
        ano = int(request.args.get('ano', date.today().year))
    except (TypeError, ValueError):
        ano = date.today().year
    meses = [f'{ano:04d}-{m:02d}' for m in range(1, 13)]
    # matriz {centro_id: {mes: valor}} do ano selecionado
    orc_map = {}
    for o in orcs:
        if (o.get('mes') or '')[:4] == str(ano):
            orc_map.setdefault(o.get('centro_custo'), {})[o['mes']] = float(o.get('valor') or 0)
    # anos com orçamento (para navegação)
    anos = sorted({int(o['mes'][:4]) for o in orcs if o.get('mes')} | {date.today().year, ano})
    return render_template('financeiro_config.html',
                           centros=centros, meses=meses, ano=ano, anos=anos, orc_map=orc_map)


@app.route('/financeiro/config/centro', methods=['POST'])
def financeiro_config_centro():
    if not _fin_config_required():
        return redirect(url_for('financeiro'))
    centros = load_centros_custo()
    cid    = request.form.get('id', '')
    codigo = request.form.get('codigo', '').strip()
    nome   = request.form.get('nome', '').strip()
    tipo   = 'OPEX' if request.form.get('tipo') == 'OPEX' else 'CAPEX'
    if not nome:
        flash('Informe o nome do centro de custo.', 'warning')
        return redirect(url_for('financeiro_config'))
    if cid:
        item = next((c for c in centros if c.get('id') == cid), None)
        if item:
            item.update({'codigo': codigo, 'nome': nome, 'tipo': tipo})
            audit_log('editar_centro_custo', cid, f'{codigo} {nome}')
    else:
        centros.append({'id': str(uuid.uuid4()), 'codigo': codigo, 'nome': nome,
                        'tipo': tipo, 'ativo': True})
        audit_log('criar_centro_custo', nome, codigo)
    save_centros_custo(centros)
    flash('Centro de custo salvo.', 'success')
    return redirect(url_for('financeiro_config'))


@app.route('/financeiro/config/centro/<cid>/excluir', methods=['POST'])
def financeiro_config_centro_excluir(cid):
    if not _fin_config_required():
        return redirect(url_for('financeiro'))
    centros = [c for c in load_centros_custo() if c.get('id') != cid]
    save_centros_custo(centros)
    save_orcamentos([o for o in load_orcamentos() if o.get('centro_custo') != cid])
    audit_log('excluir_centro_custo', cid)
    flash('Centro de custo excluído (orçamentos vinculados removidos).', 'success')
    return redirect(url_for('financeiro_config'))


@app.route('/financeiro/config/orcamento', methods=['POST'])
def financeiro_config_orcamento():
    if not _fin_config_required():
        return redirect(url_for('financeiro'))
    try:
        ano = int(request.form.get('ano', date.today().year))
    except (TypeError, ValueError):
        ano = date.today().year
    # substitui o orçamento do ano pelos valores enviados (campos "orc|<cid>|<mes>")
    orcs = [o for o in load_orcamentos() if (o.get('mes') or '')[:4] != str(ano)]
    for name, raw in request.form.items():
        if not name.startswith('orc|'):
            continue
        _, cid, mes = name.split('|', 2)
        v = parse_brl(raw)
        if v:
            orcs.append({'id': f'{cid}|{mes}', 'centro_custo': cid, 'mes': mes, 'valor': v})
    save_orcamentos(orcs)
    audit_log('salvar_orcamento', str(ano))
    flash(f'Orçamento de {ano} salvo.', 'success')
    return redirect(url_for('financeiro_config', ano=ano))


@app.route('/consolidado')
def consolidado():
    """Painel financeiro consolidado de TODAS as contratadas (perfil RUMO/master)."""
    u = current_user()
    if not (session.get('admin_ok') or (u and u.get('role') in ADMIN_ROLES)):
        flash('Acesso restrito ao painel consolidado.', 'danger')
        return redirect(url_for('dashboard'))

    registros  = load_data()
    cfg        = load_contratos_config()
    tms_cfg    = load_tms_config()

    # Universo de contratadas (config + registros)
    contratadas = sorted(
        {c.get('contratada', '') for c in cfg.values() if c.get('contratada')}
        | {r.get('contratada', '') for r in registros if r.get('contratada')}
    )

    # ── KPIs consolidados ──
    total_valor_contrato = sum(valor_efetivo(c) for c in cfg.values())
    total_medido = sum(r.get('valor_medido', 0) for r in registros)
    saldo        = total_valor_contrato - total_medido
    pct_global   = round(total_medido / total_valor_contrato * 100, 1) if total_valor_contrato else 0

    n_contratadas = len(contratadas)
    pares_contrato = {(r.get('contratada', ''), r.get('contrato', '')) for r in registros if r.get('contrato')}
    pares_contrato |= {(c.get('contratada', ''), c.get('contrato', '')) for c in cfg.values() if c.get('contrato')}
    n_contratos = len(pares_contrato)

    reg_ord = sorted(registros, key=lambda r: r.get('semana_referencia', ''))
    ultima_semana = reg_ord[-1].get('semana_referencia', '') if reg_ord else ''
    medido_semana = sum(r.get('valor_medido', 0) for r in registros if r.get('semana_referencia') == ultima_semana)

    # Medição agregada numa única passada (reuso em por_contratada / por_área / contratos)
    medido_por_par = {}         # (contratada, contrato) → medido acumulado
    medido_por_contratada = {}  # contratada → medido acumulado
    for r in registros:
        ct = r.get('contratada', '')
        co = r.get('contrato', '')
        v  = r.get('valor_medido', 0) or 0
        medido_por_par[(ct, co)] = medido_por_par.get((ct, co), 0) + v
        medido_por_contratada[ct] = medido_por_contratada.get(ct, 0) + v

    # avanço físico médio considerando o último registro de cada contrato
    ultimo_af = {}
    for r in reg_ord:
        ultimo_af[(r.get('contratada'), r.get('contrato'))] = r.get('avanco_fisico', 0)
    af_medio = round(sum(ultimo_af.values()) / len(ultimo_af), 1) if ultimo_af else 0

    # ── Curva S consolidada (acumulado semanal somando todas) ──
    semanas = {}
    for r in reg_ord:
        s = r.get('semana_referencia', '')
        semanas[s] = semanas.get(s, 0) + r.get('valor_medido', 0)
    semanas_sorted = sorted(semanas)
    curva_acum, acum = [], 0
    for s in semanas_sorted:
        acum += semanas[s]
        curva_acum.append(round(acum, 2))
    # chart_labels / curva_base / curva_fis_base são montados adiante sobre a
    # linha do tempo ESTENDIDA (semanas + meses futuros) — ver bloco Forecast.

    # ── Curva S Física consolidada (média ponderada por valor contratado) ──
    _af_hist = {}  # (contratada, contrato) → {semana: avanco_fisico}
    for r in reg_ord:
        _k = (r.get('contratada'), r.get('contrato'))
        _s = r.get('semana_referencia', '')
        if _s:
            _af_hist.setdefault(_k, {})[_s] = r.get('avanco_fisico') or 0
    _ct_val = {(c.get('contratada'), c.get('contrato')): valor_efetivo(c)
               for c in cfg.values()}
    curva_fis_acum = []
    for s in semanas_sorted:
        soma, peso = 0.0, 0.0
        for _k, _hist in _af_hist.items():
            _known = {w: v for w, v in _hist.items() if w <= s}
            if not _known:
                continue
            _af  = _known[max(_known)]
            _p   = _ct_val.get(_k, 1.0)
            soma += _af * _p
            peso += _p
        curva_fis_acum.append(round(soma / peso, 1) if peso else None)

    # Baseline física: média ponderada mensal
    _fis_m = {}
    for c in cfg.values():
        _k  = (c.get('contratada'), c.get('contrato'))
        _p  = valor_efetivo(c)
        _lb = c.get('lb_fis_planejado') or {it['semana']: it['percentual'] for it in c.get('linha_base_fisica', [])}
        for _m, _v in _lb.items():
            _fis_m.setdefault(_m, []).append((float(_v or 0), _p))
    _fis_cumul = {}
    for _m in sorted(_fis_m):
        _soma = sum(v * p for v, p in _fis_m[_m])
        _peso = sum(p for _, p in _fis_m[_m])
        _fis_cumul[_m] = round(_soma / _peso, 1) if _peso else 0
    _fis_months = sorted(_fis_cumul)
    # curva_fis_base é amostrada adiante sobre a linha do tempo estendida.

    # ── Forecast das Curvas S ────────────────────────────────────────────────
    # Estende o eixo do tempo até o fim do baseline/contrato e projeta a curva
    # como a linha de base deslocada pelo desvio atual (forecast). Ver helpers.
    future_dates = _future_dates(cfg, semanas_sorted)
    all_dates = semanas_sorted + future_dates
    chart_labels = [format_date_br(s) for s in all_dates]

    # Recalcula baselines e realizados sobre a linha do tempo estendida
    curva_base = fin_base_curve(cfg, '', all_dates)
    curva_fis_base = [
        next(((_fis_cumul[bm]) for bm in reversed(_fis_months) if bm <= _month_of_week(s)), None)
        for s in all_dates
    ]
    _nf = len(future_dates)
    curva_acum = curva_acum + [None] * _nf          # realizado não existe no futuro
    curva_fis_acum = curva_fis_acum + [None] * _nf

    curva_fc = _forecast_series(curva_acum, curva_base)
    curva_fis_fc = _forecast_series(curva_fis_acum, curva_fis_base, clamp_max=100)

    # ── Linha de base REPLANEJADA (mostrada só se houver replanejamento) ───────
    # Cada contrato pode ter replanejamentos (lb_*_replanejados = lista de {mês:valor}).
    # A baseline replanejada consolidada usa, por contrato, o replan mais recente
    # sobreposto ao planejado original; contratos sem replan mantêm o original.
    # Financeira: soma acumulada mensal. Física: média ponderada por valor de contrato.

    # Financeira replanejada
    _fin_repl_m, _has_repl_fin = {}, False
    for c in cfg.values():
        _base = c.get('lb_fin_planejado') or {it['semana']: it['valor'] for it in c.get('linha_base_financeira', [])}
        _rev, _ch = _mesclar_replan(_base, c.get('lb_fin_replanejados'))
        _has_repl_fin = _has_repl_fin or _ch
        for _m, _v in _rev.items():
            _fin_repl_m[_m] = _fin_repl_m.get(_m, 0) + float(_v or 0)
    curva_repl = None
    if _has_repl_fin:
        _cumul, _acc = {}, 0
        for _m in sorted(_fin_repl_m):
            _acc += _fin_repl_m[_m]
            _cumul[_m] = round(_acc, 2)
        _months_r = sorted(_cumul)
        curva_repl = [
            next((_cumul[bm] for bm in reversed(_months_r) if bm <= _month_of_week(s)), None)
            for s in all_dates
        ]

    # Física replanejada (média ponderada por valor de contrato)
    _fis_repl_m, _has_repl_fis = {}, False
    for c in cfg.values():
        _p = valor_efetivo(c)
        _base = c.get('lb_fis_planejado') or {it['semana']: it['percentual'] for it in c.get('linha_base_fisica', [])}
        _rev, _ch = _mesclar_replan(_base, c.get('lb_fis_replanejados'))
        _has_repl_fis = _has_repl_fis or _ch
        for _m, _v in _rev.items():
            _fis_repl_m.setdefault(_m, []).append((float(_v or 0), _p))
    curva_fis_repl = None
    if _has_repl_fis:
        _fc = {}
        for _m in sorted(_fis_repl_m):
            _soma = sum(v * p for v, p in _fis_repl_m[_m])
            _peso = sum(p for _, p in _fis_repl_m[_m])
            _fc[_m] = round(_soma / _peso, 1) if _peso else 0
        _months_r = sorted(_fc)
        curva_fis_repl = [
            next((_fc[bm] for bm in reversed(_months_r) if bm <= _month_of_week(s)), None)
            for s in all_dates
        ]

    # ── Por contratada ──
    por_contratada = []
    for ct in contratadas:
        valor  = sum(valor_efetivo(c) for c in cfg.values() if c.get('contratada') == ct)
        medido = medido_por_contratada.get(ct, 0)
        por_contratada.append({
            'contratada': ct,
            'valor':  valor,
            'medido': medido,
            'saldo':  valor - medido,
            'pct':    round(medido / valor * 100, 1) if valor else 0,
        })
    por_contratada.sort(key=lambda x: -x['medido'])

    _PALETTE = ['rgba(0,212,255,.85)', 'rgba(141,198,63,.85)', 'rgba(240,165,0,.85)',
                'rgba(255,107,122,.85)', 'rgba(155,89,182,.85)', 'rgba(52,211,153,.85)',
                'rgba(96,165,250,.85)']
    cores = [_PALETTE[i % len(_PALETTE)] for i in range(len(por_contratada))]

    # ── Por área de contrato (totais + detalhamento por contrato numa só passada) ──
    area_map = {}
    contratos_por_area: dict = {}
    for c in cfg.values():
        area     = (c.get('area_contrato') or '').strip() or 'Não classificado'
        ct_name  = c.get('contratada', '')
        contrato = c.get('contrato', '')
        valor    = valor_efetivo(c)
        medido_c = medido_por_par.get((ct_name, contrato), 0)
        a = area_map.setdefault(area, {'area': area, 'valor': 0.0, 'medido': 0.0, 'n_contratos': 0})
        a['valor']       += valor
        a['medido']      += medido_c
        a['n_contratos'] += 1
        contratos_por_area.setdefault(area, []).append({
            'contratada': ct_name,
            'contrato':   contrato,
            'valor':      valor,
            'medido':     round(medido_c, 2),
            'saldo':      round(valor - medido_c, 2),
            'pct':        round(medido_c / valor * 100, 1) if valor else 0,   # % financeiro (medido/contratado)
            'avanco_fisico': round(ultimo_af.get((ct_name, contrato), 0) or 0, 1),  # último avanço físico do contrato
        })
    for a in area_map.values():
        a['saldo'] = a['valor'] - a['medido']
        a['pct']   = round(a['medido'] / a['valor'] * 100, 1) if a['valor'] else 0
    por_area = sorted(area_map.values(), key=lambda x: -x['valor'])

    kpis = dict(
        total_valor_contrato=total_valor_contrato,
        total_medido=total_medido,
        saldo=saldo,
        pct_global=pct_global,
        medido_semana=medido_semana,
        ultima_semana=ultima_semana,
        n_contratadas=n_contratadas,
        n_contratos=n_contratos,
        af_medio=af_medio,
    )

    tms_milestones = sorted(tms_cfg.get('milestones', []), key=lambda m: m.get('mes', ''))

    # ── Histograma consolidado (última semana com dados) ──
    _ult_sem = semanas_sorted[-1] if semanas_sorted else None
    _ult_iso = format_date_to_week(_ult_sem) if _ult_sem else None
    _hr_dir = _hr_ind = 0
    for r in registros:
        if r.get('semana_referencia') == _ult_sem:
            _hr_dir += int(r.get('total_direto',   0) or 0)
            _hr_ind += int(r.get('total_indireto', 0) or 0)
    _hp_dir = _hp_ind = 0
    if _ult_iso:
        for c in cfg.values():
            for lb in c.get('linha_base_histograma', []):
                _v = int(lb.get('semanas', {}).get(_ult_iso, 0) or 0)
                if lb.get('tipo') == 'indireto':
                    _hp_ind += _v
                else:
                    _hp_dir += _v
    _tms_h = int(tms_cfg.get('headcount_tms', 0) or 0)
    hist_kpis = {
        'semana':    _ult_iso or '',
        'mod_plan':  _hp_dir,  'mod_real':  _hr_dir,  'mod_delta':  _hr_dir - _hp_dir,
        'moi_plan':  _hp_ind,  'moi_real':  _hr_ind,  'moi_delta':  _hr_ind - _hp_ind,
        'tot_plan':  _hp_dir + _hp_ind,
        'tot_real':  _hr_dir + _hr_ind,
        'tot_delta': (_hr_dir + _hr_ind) - (_hp_dir + _hp_ind),
        'tms_plan':  _tms_h,   'tms_real':  _tms_h,   'tms_delta':  0,
    }

    # ── Suprimento stats para CONSOLIDADO ─────────────────────────────────
    _sups = load_suprimentos()
    sup_total = len(_sups)
    _sup_etapa_labels = [SUPRIMENTO_STAGE_LABELS[s] for s in SUPRIMENTO_STAGES]
    _sup_etapa_vals   = [sum(1 for x in _sups if x.get('status') == s) for s in SUPRIMENTO_STAGES]
    _sup_etapa_cores  = [SUPRIMENTO_STAGE_COLORS[s] for s in SUPRIMENTO_STAGES]
    _sup_area_map = {}
    for x in _sups:
        a = x.get('area_contrato') or 'Não classificado'
        _sup_area_map[a] = _sup_area_map.get(a, 0) + 1
    sup_valor_total = sum(x.get('valor_estimado') or 0 for x in _sups)
    sup_etapa_rows  = list(zip(_sup_etapa_labels, _sup_etapa_vals, _sup_etapa_cores))

    # Dias que o processo está na etapa atual (desde a entrada nela, via histórico)
    _hoje = datetime.now().date()

    def _dias_na_etapa(x):
        st = x.get('status')
        arrival = ''
        for h in (x.get('historico') or []):
            if h.get('status') == st and h.get('data'):
                arrival = h['data']   # mantém a entrada mais recente nessa etapa
        if not arrival:
            arrival = x.get('criado_em') or x.get('data_inicio') or ''
        arrival = (arrival or '')[:10]
        if not arrival:
            return None
        try:
            d0 = datetime.strptime(arrival, '%Y-%m-%d').date()
            return max((_hoje - d0).days, 0)
        except Exception:
            return None

    # Processos agrupados por etapa — usado no card de detalhe ao clicar no funil
    _sup_por_etapa = {s: [] for s in SUPRIMENTO_STAGES}
    for x in _sups:
        st = x.get('status')
        if st in _sup_por_etapa:
            _sup_por_etapa[st].append({
                'descricao':          x.get('descricao', ''),
                'objeto':             x.get('objeto', ''),
                'contratada':         x.get('contratada', ''),
                'valor_estimado':     x.get('valor_estimado') or 0,
                'data_inicio':        x.get('data_inicio', ''),
                'data_prev_contrato': x.get('data_prev_contrato', ''),
                'status':             st,
                'prioridade':         x.get('prioridade', ''),
                'responsavel':        x.get('responsavel', ''),
                'dias_etapa':         _dias_na_etapa(x),
            })
    _sup_stage_labels = SUPRIMENTO_STAGE_LABELS
    _sup_stage_cores  = SUPRIMENTO_STAGE_COLORS

    return render_template('consolidado.html',
                           kpis=kpis,
                           por_contratada=por_contratada,
                           por_area=por_area,
                           contratos_por_area=contratos_por_area,
                           tms_cfg=tms_cfg,
                           tms_milestones=tms_milestones,
                           hist_kpis=hist_kpis,
                           chart_labels=json.dumps(chart_labels),
                           curva_acum=json.dumps(curva_acum),
                           curva_base=json.dumps(curva_base),
                           curva_fc=json.dumps(curva_fc),
                           curva_repl=json.dumps(curva_repl),
                           curva_fis_acum=json.dumps(curva_fis_acum),
                           curva_fis_base=json.dumps(curva_fis_base),
                           curva_fis_fc=json.dumps(curva_fis_fc),
                           curva_fis_repl=json.dumps(curva_fis_repl),
                           bar_labels=json.dumps([x['contratada'] for x in por_contratada]),
                           bar_medido=json.dumps([round(x['medido'], 2) for x in por_contratada]),
                           bar_pct=json.dumps([x['pct'] for x in por_contratada]),
                           cores=json.dumps(cores),
                           sup_total=sup_total,
                           sup_valor_total=sup_valor_total,
                           sup_etapa_rows=sup_etapa_rows,
                           sup_por_etapa_labels=json.dumps(_sup_etapa_labels),
                           sup_por_etapa_vals=json.dumps(_sup_etapa_vals),
                           sup_por_etapa_cores=json.dumps(_sup_etapa_cores),
                           sup_etapa_keys=json.dumps(SUPRIMENTO_STAGES),
                           sup_por_etapa_data=json.dumps(_sup_por_etapa),
                           sup_stage_labels=json.dumps(_sup_stage_labels),
                           sup_stage_cores=json.dumps(_sup_stage_cores),
                           sup_por_area_labels=json.dumps(list(_sup_area_map.keys())),
                           sup_por_area_vals=json.dumps(list(_sup_area_map.values())))


@app.route('/construcao')
def construcao():
    return render_template('construcao.html')


@app.route('/mobilizacao')
def mobilizacao():
    u = current_user()
    if not u:
        return redirect(url_for('login'))
    mob = load_mobilizacao()

    # ── Documentações ──
    doc_counts = {'entregue': 0, 'pendente': 0, 'nao': 0, 'na': 0, 'vazio': 0}
    for d in mob['documentacoes']:
        s = d.get('status') or 'vazio'
        doc_counts[s if s in doc_counts else 'vazio'] += 1
    docs_aplicaveis = len(mob['documentacoes']) - doc_counts['na']
    docs_pct = round(doc_counts['entregue'] / docs_aplicaveis * 100) if docs_aplicaveis else 0

    # ── SSMA ──
    ssma_counts = {'entregue': 0, 'revisado': 0, 'nao': 0, 'vazio': 0}
    for d in mob['ssma']:
        s = d.get('status') or 'vazio'
        ssma_counts[s if s in ssma_counts else 'vazio'] += 1
    ssma_pct = round((ssma_counts['entregue'] + ssma_counts['revisado']) / len(mob['ssma']) * 100) if mob['ssma'] else 0

    # ── MO ──
    mo_etapas  = mob['mo']['etapas']
    mo_prev    = mo_etapas[0] if mo_etapas else {}
    mo_stages  = mo_etapas[1:] if len(mo_etapas) > 1 else []
    mo_has_data = any(e.get('geral') is not None for e in mo_etapas)
    mo_contratados = mo_etapas[2].get('geral') if len(mo_etapas) > 2 else None

    # ── Equipamentos ──
    eq_etapas  = mob['equipamentos']['etapas']
    eq_prev    = eq_etapas[0] if eq_etapas else {}
    eq_stages  = eq_etapas[1:] if len(eq_etapas) > 1 else []
    eq_has_data = any(e.get('geral') is not None for e in eq_etapas)

    # ── Canteiro ──
    canteiro_items = []
    cant_total = 0.0
    for it in mob['canteiro']:
        peso = it.get('peso', 0) or 0
        crits = it.get('criterios') or []
        av = min(1.0, sum(cr['peso'] for cr in crits if cr.get('concluido'))) if crits else 0.0
        canteiro_items.append({'item': it['item'], 'avanco': round(av * 100), 'peso': peso})
        cant_total += peso * av
    cant_total_pct = round(cant_total * 100)

    # ── Subcontratadas ──
    sub_items = []
    sub_total = 0.0
    for emp in mob['subcontratadas']:
        peso = emp.get('peso', 0) or 0
        crits = emp.get('criterios') or []
        av = min(1.0, sum(cr['peso'] for cr in crits if cr.get('concluido')))
        sub_items.append({'empresa': emp['empresa'], 'avanco': round(av * 100), 'peso': peso})
        sub_total += peso * av
    sub_total_pct = round(sub_total * 100)

    return render_template('mobilizacao.html', current_user=u,
        doc_counts=doc_counts, docs_pct=docs_pct,
        ssma_counts=ssma_counts, ssma_pct=ssma_pct,
        mo_prev=mo_prev, mo_stages=mo_stages, mo_has_data=mo_has_data,
        mo_contratados=mo_contratados,
        eq_prev=eq_prev, eq_stages=eq_stages, eq_has_data=eq_has_data,
        canteiro_items=canteiro_items, cant_total_pct=cant_total_pct,
        sub_items=sub_items, sub_total_pct=sub_total_pct,
        mob=mob,
    )


_MOB_DEFAULT = {
    "documentacoes": [
        {"item": 1,  "documento": "Anotação de Responsabilidade Técnica (ART — Obras de engenharia / serviços)", "area": "Saúde e Segurança Ocupacional", "status": None, "prazo": None},
        {"item": 2,  "documento": "Carta de Apresentação do Engenheiro Residente (Obras)", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 3,  "documento": "Garantias de Execução / Adiantamento Contratual (Seguros e Garantias)", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 4,  "documento": "Inscrição no INSS (CEI — aplicável para Obras Civis)", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 5,  "documento": "Inscrição na Prefeitura Municipal para Recolhimento de ISSQN", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 6,  "documento": "Certidão Negativa de Débito atualizada com o INSS", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 7,  "documento": "Certificado de Regularidade com o FGTS", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 8,  "documento": "Certidão Negativa de Débito com o FGTS (CRF)", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 9,  "documento": "Certidão Negativa de Débito do ISSQN com o Município", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 10, "documento": "Certidão Negativa de Débito da Receita Federal", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 11, "documento": "Certidão Negativa de Débito da Receita Estadual", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 12, "documento": "Cartão do CNPJ", "area": "Gestão de Contratos", "status": None, "prazo": None},
        {"item": 13, "documento": "Mapeamentos de Pessoas RAC", "area": "Saúde e Segurança Ocupacional", "status": None, "prazo": None},
        {"item": 14, "documento": "PRE (Programa de Ergonomia)", "area": "Saúde e Segurança Ocupacional", "status": None, "prazo": None},
        {"item": 15, "documento": "CTF — Cadastro Técnico Federal", "area": "Meio Ambiente", "status": None, "prazo": None},
        {"item": 16, "documento": "PAE (Plano de Atendimento à Emergência)", "area": "Saúde e Segurança Ocupacional", "status": None, "prazo": None},
        {"item": 17, "documento": "APR (Análise Preliminar de Risco — Makro)", "area": "Saúde e Segurança Ocupacional", "status": None, "prazo": None},
    ],
    "ssma": [
        {"item": 1, "documento": "APR HO", "area": "Saúde e Segurança Ocupacional", "status": None},
        {"item": 2, "documento": "PCMSO", "area": "Saúde e Segurança Ocupacional", "status": None},
        {"item": 3, "documento": "PCMAT (aplicável na construção civil)", "area": "Saúde e Segurança Ocupacional", "status": None},
        {"item": 4, "documento": "PGR", "area": "Saúde e Segurança Ocupacional", "status": None},
    ],
    "mo": {
        "etapas": [
            {"etapa": "Previsto Contratual até término MOB", "geral": None, "mod": None, "moi": None},
            {"etapa": "Em contratação",                     "geral": None, "mod": None, "moi": None},
            {"etapa": "Contratados",                        "geral": None, "mod": None, "moi": None},
            {"etapa": "Exames Prontos",                     "geral": None, "mod": None, "moi": None},
            {"etapa": "Treinamentos (RACs) Executados",     "geral": None, "mod": None, "moi": None},
            {"etapa": "Documentação entregue",              "geral": None, "mod": None, "moi": None},
            {"etapa": "Integração Concluída",               "geral": None, "mod": None, "moi": None},
            {"etapa": "Crachá Disponível",                  "geral": None, "mod": None, "moi": None},
        ],
        "comentarios": ""
    },
    "equipamentos": {
        "etapas": [
            {"etapa": "Previsto Contratual até término MOB", "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Em mobilização",                      "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Mobilizados (disponibilizados)",      "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Pré-comissionados",                   "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Documentação entregue",               "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Em Barão de Cocais",                  "geral": None, "grande_porte": None, "pequeno_porte": None},
            {"etapa": "Comissionados",                       "geral": None, "grande_porte": None, "pequeno_porte": None},
        ],
        "comentarios": ""
    },
    "canteiro": [
        {"item": "Layout canteiro",                                          "peso": 0.05, "criterios": [{"descricao": "Entregue", "peso": 0.5, "concluido": False}, {"descricao": "Aprovado", "peso": 0.5, "concluido": False}]},
        {"item": "Montagem e construção do canteiro (provisório)",           "peso": 0.10, "criterios": [{"descricao": "Instalação Container", "peso": 0.33, "concluido": False}, {"descricao": "Instalação Banheiro Químico", "peso": 0.33, "concluido": False}, {"descricao": "Instalação Tenda", "peso": 0.33, "concluido": False}]},
        {"item": "Montagem e construção do canteiro (definitivo)",           "peso": 0.00, "criterios": []},
        {"item": "Mobilização de equipamentos para canteiro",                "peso": 0.25, "criterios": [{"descricao": "Caminhão Munck", "peso": 0.3, "concluido": False}, {"descricao": "Container", "peso": 0.4, "concluido": False}, {"descricao": "Gerador", "peso": 0.3, "concluido": False}]},
        {"item": "Banheiro — Instalações sanitárias de canteiro",            "peso": 0.05, "criterios": [{"descricao": "Chegada no site", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Energia/TI — Rebaixamento e instalações elétricas",        "peso": 0.10, "criterios": [{"descricao": "Início das atividades", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Água / Esgoto — Execução de instal. de água e esgoto",     "peso": 0.10, "criterios": [{"descricao": "Início das atividades", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Incêndio — Instalações de sistema de incêndio",            "peso": 0.05, "criterios": [{"descricao": "Início das atividades", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Fossa — Execução de fossa / tratamento de esgoto",         "peso": 0.05, "criterios": [{"descricao": "Início das atividades", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Refeição — Execução de refeitório",                        "peso": 0.15, "criterios": [{"descricao": "Chegada no site", "peso": 0.3, "concluido": False}, {"descricao": "Operacional", "peso": 0.7, "concluido": False}]},
        {"item": "Aprovação",                                                "peso": 0.10, "criterios": [{"descricao": "Vistoria pela SSMA", "peso": 0.2, "concluido": False}, {"descricao": "Aprovação pela Gestão do Contrato", "peso": 0.4, "concluido": False}, {"descricao": "Aceite pela SSMA", "peso": 0.4, "concluido": False}]},
    ],
    "subcontratadas": [
        {"empresa": "Container",               "peso": 0.20, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.2, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.4, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.4, "concluido": False}]},
        {"empresa": "Banheiro Químico",        "peso": 0.10, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.2, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.4, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.4, "concluido": False}]},
        {"empresa": "Locação Veículos",        "peso": 0.10, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.2, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.4, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.4, "concluido": False}]},
        {"empresa": "Transporte de Mão de Obra","peso": 0.20, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.2, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.1, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.3, "concluido": False}, {"descricao": "Crachás liberados", "peso": 0.4, "concluido": False}]},
        {"empresa": "Alimentação",             "peso": 0.20, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.2, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.4, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.4, "concluido": False}]},
        {"empresa": "Terraplenagem",           "peso": 0.20, "criterios": [{"descricao": "Documentação entregue — cédula de contrato para Subcontratação", "peso": 0.1, "concluido": False}, {"descricao": "Documentação — aprovação cédula de contrato", "peso": 0.3, "concluido": False}, {"descricao": "Contratação realizada / evidenciada", "peso": 0.3, "concluido": False}, {"descricao": "Crachás liberados", "peso": 0.3, "concluido": False}]},
    ],
}


def load_mobilizacao():
    if os.path.exists(MOBILIZACAO_FILE):
        try:
            with open(MOBILIZACAO_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return json.loads(json.dumps(_MOB_DEFAULT))


def save_mobilizacao(data):
    with open(MOBILIZACAO_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.route('/mobilizacao/config')
def mobilizacao_config():
    u = current_user()
    if not u:
        return redirect(url_for('login'))
    mob = load_mobilizacao()
    return render_template('mobilizacao_config.html', current_user=u, mob=mob)


def validar_datas_bl(inicio_bl, fim_bl, real, forecast):
    """Regra BL: SE Real > Fim BL → Forecast obrigatório.
    Retorna (True, None) se válido ou (False, mensagem) se inválido."""
    if real and fim_bl:
        try:
            from datetime import date as _d
            if _d.fromisoformat(real) > _d.fromisoformat(fim_bl) and not (forecast or '').strip():
                return False, 'Forecast obrigatório quando Real > Fim BL'
        except ValueError:
            return False, 'Data inválida no grupo BL'
    return True, None


@app.route('/mobilizacao/documentos')
def mobilizacao_documentos():
    u = current_user()
    if not u:
        return redirect(url_for('login'))
    mob = load_mobilizacao()
    docs = []
    for d in mob.get('documentacoes', []):
        docs.append({
            'item':      d.get('item', ''),
            'documento': d.get('documento', ''),
            'area':      d.get('area', ''),
            'tipo':      'Contrato',
            'status':    d.get('status') or '',
            'prazo':     d.get('prazo') or '',
            'url':       d.get('url') or '',
            'inicio_bl': d.get('inicio_bl') or '',
            'fim_bl':    d.get('fim_bl') or '',
            'real':      d.get('real') or '',
            'forecast':  d.get('forecast') or '',
        })
    for d in mob.get('ssma', []):
        docs.append({
            'item':      d.get('item', ''),
            'documento': d.get('documento', ''),
            'area':      d.get('area', ''),
            'tipo':      'SSMA',
            'status':    d.get('status') or '',
            'prazo':     '',
            'url':       d.get('url') or '',
            'inicio_bl': d.get('inicio_bl') or '',
            'fim_bl':    d.get('fim_bl') or '',
            'real':      d.get('real') or '',
            'forecast':  d.get('forecast') or '',
        })
    return render_template('mobilizacao_documentos.html', current_user=u, docs=docs)


@app.route('/mobilizacao/config/salvar', methods=['POST'])
def mobilizacao_config_salvar():
    u = current_user()
    if not u:
        return jsonify({'ok': False, 'erro': 'Não autenticado'}), 401
    payload = request.get_json(force=True)
    if not payload or 'secao' not in payload:
        return jsonify({'ok': False, 'erro': 'Payload inválido'}), 400
    secao = payload['secao']

    # ── Validação BL: determina quais itens verificar por seção ──
    if secao == 'documentacoes':
        itens_bl = payload.get('documentacoes', [])
    elif secao == 'ssma':
        itens_bl = payload.get('ssma', [])
    elif secao == 'mo':
        itens_bl = payload.get('mo', {}).get('etapas', [])
    elif secao == 'equipamentos':
        itens_bl = payload.get('equipamentos', {}).get('etapas', [])
    elif secao == 'canteiro':
        itens_bl = payload.get('canteiro', [])
    elif secao == 'subcontratadas':
        itens_bl = payload.get('subcontratadas', [])
    else:
        itens_bl = []

    for item in itens_bl:
        nome = (item.get('documento') or item.get('etapa') or
                item.get('item') or item.get('empresa') or 'item')
        ok, err = validar_datas_bl(
            item.get('inicio_bl'), item.get('fim_bl'),
            item.get('real'),      item.get('forecast')
        )
        if not ok:
            return jsonify({'ok': False, 'erro': f'"{nome}": {err}'}), 422

    mob = load_mobilizacao()
    if secao in mob and secao in payload:
        mob[secao] = payload[secao]
    save_mobilizacao(mob)
    return jsonify({'ok': True})


@app.route('/mobilizacao/config/exportar-excel')
def mobilizacao_exportar_excel():
    u = current_user()
    if not u:
        return redirect(url_for('login'))
    mob = load_mobilizacao()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io
        wb = Workbook()
        hdr_font  = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
        hdr_fill  = PatternFill('solid', fgColor='FFFF7043')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        def _sw(ws, hdrs, widths):
            ws.append(hdrs)
            for i, cell in enumerate(ws[1]):
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
                ws.column_dimensions[get_column_letter(i+1)].width = widths[i]
        # Sheet 1 — Documentações
        ws = wb.active; ws.title = 'Documentações'
        _sw(ws, ['Item','Documento','Área','Status','Prazo','URL','Início BL','Fim BL','Real','Forecast'],
                [6,50,22,15,12,35,12,12,12,12])
        for d in mob.get('documentacoes', []):
            ws.append([d.get('item'),d.get('documento'),d.get('area'),d.get('status'),
                       d.get('prazo'),d.get('url'),d.get('inicio_bl'),d.get('fim_bl'),
                       d.get('real'),d.get('forecast')])
        # Sheet 2 — SSMA
        ws2 = wb.create_sheet('SSMA')
        _sw(ws2, ['Item','Documento','Área','Status','URL','Início BL','Fim BL','Real','Forecast'],
                 [6,45,22,15,35,12,12,12,12])
        for d in mob.get('ssma', []):
            ws2.append([d.get('item'),d.get('documento'),d.get('area'),d.get('status'),
                        d.get('url'),d.get('inicio_bl'),d.get('fim_bl'),d.get('real'),d.get('forecast')])
        # Sheet 3 — MO
        ws3 = wb.create_sheet('MO')
        _sw(ws3, ['Etapa','Geral','MOD','MOI','Início BL','Fim BL','Real','Forecast'],
                 [38,8,8,8,12,12,12,12])
        for e in mob.get('mo', {}).get('etapas', []):
            ws3.append([e.get('etapa'),e.get('geral'),e.get('mod'),e.get('moi'),
                        e.get('inicio_bl'),e.get('fim_bl'),e.get('real'),e.get('forecast')])
        # Sheet 4 — Equipamentos
        ws4 = wb.create_sheet('Equipamentos')
        _sw(ws4, ['Etapa','Geral','Grande Porte','Pequeno Porte','Início BL','Fim BL','Real','Forecast'],
                 [38,8,13,13,12,12,12,12])
        for e in mob.get('equipamentos', {}).get('etapas', []):
            ws4.append([e.get('etapa'),e.get('geral'),e.get('grande_porte'),e.get('pequeno_porte'),
                        e.get('inicio_bl'),e.get('fim_bl'),e.get('real'),e.get('forecast')])
        # Sheet 5 — Canteiro
        ws5 = wb.create_sheet('Canteiro')
        _sw(ws5, ['Item','Peso (%)','Critérios (sep. ;)','Pesos Critérios (sep. ;)','Concluídos 0/1 (sep. ;)','Início BL','Fim BL','Real','Forecast'],
                 [38,10,50,25,25,12,12,12,12])
        for it in mob.get('canteiro', []):
            crits = it.get('criterios', [])
            ws5.append([it.get('item'), round(it.get('peso', 0)*100, 1),
                        ';'.join(c.get('descricao', '') for c in crits),
                        ';'.join(str(c.get('peso', 0)) for c in crits),
                        ';'.join('1' if c.get('concluido') else '0' for c in crits),
                        it.get('inicio_bl'), it.get('fim_bl'), it.get('real'), it.get('forecast')])
        # Sheet 6 — Subcontratadas
        ws6 = wb.create_sheet('Subcontratadas')
        _sw(ws6, ['Empresa','Peso (%)','Critérios (sep. ;)','Pesos Critérios (sep. ;)','Concluídos 0/1 (sep. ;)','Início BL','Fim BL','Real','Forecast'],
                 [28,10,50,25,25,12,12,12,12])
        for emp in mob.get('subcontratadas', []):
            crits = emp.get('criterios', [])
            ws6.append([emp.get('empresa'), round(emp.get('peso', 0)*100, 1),
                        ';'.join(c.get('descricao', '') for c in crits),
                        ';'.join(str(c.get('peso', 0)) for c in crits),
                        ';'.join('1' if c.get('concluido') else '0' for c in crits),
                        emp.get('inicio_bl'), emp.get('fim_bl'), emp.get('real'), emp.get('forecast')])
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        from flask import send_file
        return send_file(buf, as_attachment=True,
                         download_name='mobilizacao_modelo.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'Erro ao gerar Excel: {e}', 'danger')
        return redirect(url_for('mobilizacao_config'))


@app.route('/mobilizacao/config/importar-excel', methods=['POST'])
def mobilizacao_importar_excel():
    u = current_user()
    if not u:
        return jsonify({'ok': False, 'erro': 'Não autenticado'}), 401
    if 'arquivo' not in request.files:
        return jsonify({'ok': False, 'erro': 'Nenhum arquivo enviado'}), 400
    f = request.files['arquivo']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'ok': False, 'erro': 'Formato inválido — envie .xlsx'}), 400
    try:
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(f.read()), data_only=True)
        mob = load_mobilizacao()
        def sv(v):
            if v is None: return None
            s = str(v).strip(); return s or None
        def dt(v):
            if v is None: return None
            if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
            s = str(v).strip(); return s or None
        def ni(v):
            try: return int(float(v)) if v is not None else None
            except: return None
        def nf(v):
            try: return float(v) if v is not None else None
            except: return None
        if 'Documentações' in wb.sheetnames:
            docs = []
            for i, row in enumerate(wb['Documentações'].iter_rows(min_row=2, values_only=True)):
                if not row[1]: continue
                docs.append({'item': ni(row[0]) or i+1, 'documento': sv(row[1]) or '',
                    'area': sv(row[2]) or 'Gestao de Contratos', 'status': sv(row[3]),
                    'prazo': dt(row[4]), 'url': sv(row[5]),
                    'inicio_bl': dt(row[6]), 'fim_bl': dt(row[7]), 'real': dt(row[8]), 'forecast': dt(row[9])})
            if docs: mob['documentacoes'] = docs
        if 'SSMA' in wb.sheetnames:
            ssma = []
            for i, row in enumerate(wb['SSMA'].iter_rows(min_row=2, values_only=True)):
                if not row[1]: continue
                ssma.append({'item': ni(row[0]) or i+1, 'documento': sv(row[1]) or '',
                    'area': sv(row[2]) or 'Saude e Seguranca Ocupacional', 'status': sv(row[3]),
                    'url': sv(row[4]),
                    'inicio_bl': dt(row[5]), 'fim_bl': dt(row[6]), 'real': dt(row[7]), 'forecast': dt(row[8])})
            if ssma: mob['ssma'] = ssma
        if 'MO' in wb.sheetnames:
            etapas = []
            for row in wb['MO'].iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                etapas.append({'etapa': sv(row[0]) or '', 'geral': ni(row[1]), 'mod': ni(row[2]), 'moi': ni(row[3]),
                    'inicio_bl': dt(row[4]), 'fim_bl': dt(row[5]), 'real': dt(row[6]), 'forecast': dt(row[7])})
            if etapas: mob['mo']['etapas'] = etapas
        if 'Equipamentos' in wb.sheetnames:
            etapas = []
            for row in wb['Equipamentos'].iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                etapas.append({'etapa': sv(row[0]) or '', 'geral': ni(row[1]), 'grande_porte': ni(row[2]), 'pequeno_porte': ni(row[3]),
                    'inicio_bl': dt(row[4]), 'fim_bl': dt(row[5]), 'real': dt(row[6]), 'forecast': dt(row[7])})
            if etapas: mob['equipamentos']['etapas'] = etapas
        if 'Canteiro' in wb.sheetnames:
            cant = []
            for row in wb['Canteiro'].iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                pp = nf(row[1]) or 0
                descs = [x.strip() for x in (sv(row[2]) or '').split(';') if x.strip()]
                pesos = [x.strip() for x in (sv(row[3]) or '').split(';') if x.strip()]
                concs = [x.strip() for x in (sv(row[4]) or '').split(';') if x.strip()]
                crit = [{'descricao': d, 'peso': float(pesos[i]) if i < len(pesos) else 0,
                         'concluido': concs[i] == '1' if i < len(concs) else False}
                        for i, d in enumerate(descs)]
                cant.append({'item': sv(row[0]) or '', 'peso': round(pp/100, 4), 'criterios': crit,
                    'inicio_bl': dt(row[5]), 'fim_bl': dt(row[6]), 'real': dt(row[7]), 'forecast': dt(row[8])})
            if cant: mob['canteiro'] = cant
        if 'Subcontratadas' in wb.sheetnames:
            subs = []
            for row in wb['Subcontratadas'].iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                pp = nf(row[1]) or 0
                descs = [x.strip() for x in (sv(row[2]) or '').split(';') if x.strip()]
                pesos = [x.strip() for x in (sv(row[3]) or '').split(';') if x.strip()]
                concs = [x.strip() for x in (sv(row[4]) or '').split(';') if x.strip()]
                crit = [{'descricao': d, 'peso': float(pesos[i]) if i < len(pesos) else 0,
                         'concluido': concs[i] == '1' if i < len(concs) else False}
                        for i, d in enumerate(descs)]
                subs.append({'empresa': sv(row[0]) or '', 'peso': round(pp/100, 4), 'criterios': crit,
                    'inicio_bl': dt(row[5]), 'fim_bl': dt(row[6]), 'real': dt(row[7]), 'forecast': dt(row[8])})
            if subs: mob['subcontratadas'] = subs
        save_mobilizacao(mob)
        return jsonify({'ok': True, 'msg': 'Importado com sucesso!'})
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao processar arquivo: {str(e)}'}), 500


@app.route('/mas-dashboard')
def mas_dashboard():
    return render_template('MAS_TMS_Dashboard_18Jun (1).html')


@app.route('/registros')
def index():
    todos = load_data()
    todas_contratadas = sorted(set(r.get('contratada', '') for r in todos if r.get('contratada')))

    filtro_contratada = request.args.get('contratada', '')
    filtro_semana = request.args.get('semana', '')

    # Usuário vinculado a uma contratada: vê apenas os dados dela
    vc = viewer_contratada()
    if vc:
        filtro_contratada = vc
        todas_contratadas = [vc]

    registros = todos
    if filtro_contratada:
        registros = [r for r in registros if filtro_contratada.lower() in r.get('contratada', '').lower()]
    if filtro_semana:
        registros = [r for r in registros if r.get('semana_referencia') == filtro_semana]

    registros = scope_registros(registros)  # restringe ao contrato do usuário
    registros.sort(key=lambda r: r.get('semana_referencia', ''), reverse=True)

    return render_template('index.html',
                           registros=registros,
                           contratadas=todas_contratadas,
                           filtro_contratada=filtro_contratada,
                           filtro_semana=filtro_semana)


@app.route('/novo', methods=['GET', 'POST'])
def novo():
    if not can_create():
        flash('Você não tem permissão para criar registros.', 'warning')
        return redirect(url_for('index'))
    if request.method == 'POST':
        contratada = request.form.get('contratada', '').strip()
        contratada_custom = request.form.get('contratada_custom', '').strip()
        if contratada == '__outro__' and contratada_custom:
            contratada = contratada_custom

        semana_input = request.form.get('semana_referencia', '')
        semana_referencia = get_monday(semana_input)
        contrato = request.form.get('contrato', '').strip()

        # Usuário vinculado a um contrato só cria registros do próprio escopo
        _vc, _vk = viewer_contratada(), viewer_contrato()
        if _vc:
            contratada = _vc
        if _vk:
            contrato = _vk

        trabalhos_notaveis = request.form.get('trabalhos_notaveis', '').strip()
        pontos_atencao = request.form.get('pontos_atencao', '').strip()
        valor_medido_str = request.form.get('valor_medido', '0').strip().replace(',', '.')
        avanco_fisico_str = request.form.get('avanco_fisico', '0').strip().replace(',', '.')
        pluviometria = parse_pluviometria(request.form)

        erros = []
        if not contratada:
            erros.append('Contratada é obrigatória.')
        if not semana_referencia:
            erros.append('Semana de referência é obrigatória.')
        if not contrato:
            erros.append('Contrato é obrigatório.')
        if not trabalhos_notaveis:
            erros.append('Trabalhos notáveis são obrigatórios.')

        try:
            valor_medido = float(valor_medido_str) if valor_medido_str else 0.0
        except:
            valor_medido = 0.0
            erros.append('Valor medido inválido.')

        try:
            avanco_fisico = float(avanco_fisico_str) if avanco_fisico_str else 0.0
        except:
            avanco_fisico = 0.0

        registros = load_data()
        for r in registros:
            if r['contratada'] == contratada and r['semana_referencia'] == semana_referencia:
                erros.append(f'Já existe um registro para "{contratada}" na semana de {format_date_br(semana_referencia)}.')
                break

        efetivo, total_direto, total_indireto = parse_efetivo(request.form)
        equipamentos = parse_equipamentos(request.form)
        acoes_realizadas        = parse_acoes_realizadas(request.form)
        acoes_justificativas    = _parse_json_dict(request.form, 'acoes_justificativas_json')
        acoes_forecast          = _parse_json_dict(request.form, 'acoes_forecast_json')
        equipamentos_realizados = _parse_json_dict(request.form, 'equipamentos_realizados_json')
        histograma_realizados   = _parse_json_dict(request.form, 'histograma_realizados_json')
        avanco_fisico_fc        = _parse_json_dict(request.form, 'avanco_fisico_forecast_json')

        if erros:
            for e in erros:
                flash(e, 'danger')
            return render_template('form.html',
                                   modo='novo',
                                   contratadas=get_contratadas(),
                                   pluviometria_opcoes=PLUVIOMETRIA_OPCOES,
                                   dias_semana=DIAS_SEMANA,
                                   form_data=request.form,
                                   pluviometria_data=pluviometria,
                                   acoes_realizadas=acoes_realizadas,
                                   acoes_justificativas=acoes_justificativas,
                                   avanco_fisico_forecast=avanco_fisico_fc,
                                   avanco_fisico_fc_manual=avanco_fisico_fc_manual_map(),
                                   equipamentos_realizados=equipamentos_realizados,
                                   histograma_realizados=histograma_realizados,
                                   contratos_cfg_json=json.dumps(load_contratos_config()))

        _bl_ok_n, _bl_err_n = validar_datas_bl(
            request.form.get('inicio_bl','').strip() or None,
            request.form.get('fim_bl','').strip() or None,
            request.form.get('data_real','').strip() or None,
            request.form.get('data_forecast','').strip() or None,
        )
        if not _bl_ok_n:
            flash(f'Datas BL: {_bl_err_n}', 'danger')
            return redirect(request.url)

        novo_registro = {
            'id': str(uuid.uuid4()),
            'contrato': contrato,
            'contratada': contratada,
            'semana_referencia': semana_referencia,
            'trabalhos_notaveis': trabalhos_notaveis,
            'efetivo': efetivo,
            'total_direto': total_direto,
            'total_indireto': total_indireto,
            'equipamentos': equipamentos,
            'pontos_atencao': pontos_atencao,
            'valor_medido': valor_medido,
            'avanco_fisico': avanco_fisico,
            'avanco_fisico_forecast': avanco_fisico_fc,
            'pluviometria': pluviometria,
            'acoes_realizadas': acoes_realizadas,
            'acoes_justificativas': acoes_justificativas,
            'equipamentos_realizados': equipamentos_realizados,
            'histograma_realizados': histograma_realizados,
            'inicio_bl':     request.form.get('inicio_bl','').strip() or None,
            'fim_bl':        request.form.get('fim_bl','').strip() or None,
            'data_real':     request.form.get('data_real','').strip() or None,
            'data_forecast': request.form.get('data_forecast','').strip() or None,
            'criado_em': datetime.now().isoformat(),
            'atualizado_em': datetime.now().isoformat(),
            'criado_por': current_user_label(),
        }

        registros.append(novo_registro)
        save_data(registros)
        recompute_forecast_acoes(contratada, contrato, overrides=acoes_forecast)
        audit_log('criar_registro', novo_registro['id'],
                  f'{contratada} / {contrato} / {format_date_br(semana_referencia)}')
        flash(f'Registro de "{contratada}" para a semana de {format_date_br(semana_referencia)} salvo com sucesso!', 'success')
        return redirect(url_for('index'))

    # Pré-seleciona contratada/contrato do usuário vinculado a um contrato
    _vc, _vk = viewer_contratada(), viewer_contrato()
    _form_inicial = {}
    if _vc:
        _form_inicial = {'contratada': _vc, 'contrato': _vk or ''}
    return render_template('form.html',
                           modo='novo',
                           contratadas=([_vc] if _vc else get_contratadas()),
                           pluviometria_opcoes=PLUVIOMETRIA_OPCOES,
                           dias_semana=DIAS_SEMANA,
                           form_data=_form_inicial,
                           pluviometria_data={},
                           acoes_realizadas={},
                           acoes_justificativas={},
                           avanco_fisico_forecast={},
                           avanco_fisico_fc_manual=avanco_fisico_fc_manual_map(),
                           acoes_acumulado=acoes_acumulado_map(),
                           acoes_real_sem=acoes_real_semanas_map(),
                           acoes_fc_manual=acoes_forecast_manual_map(),
                           equipamentos_realizados={},
                           histograma_realizados={},
                           contratos_cfg_json=json.dumps(load_contratos_config()))


@app.route('/editar/<id>', methods=['GET', 'POST'])
def editar(id):
    if not can_write():
        flash('Você não tem permissão para editar registros.', 'warning')
        return redirect(url_for('index'))
    registros = load_data()
    registro = next((r for r in registros if r['id'] == id), None)

    if not registro:
        flash('Registro não encontrado.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        contratada = request.form.get('contratada', '').strip()
        contratada_custom = request.form.get('contratada_custom', '').strip()
        if contratada == '__outro__' and contratada_custom:
            contratada = contratada_custom

        semana_input = request.form.get('semana_referencia', '')
        semana_referencia = get_monday(semana_input)
        contrato = request.form.get('contrato', '').strip()
        trabalhos_notaveis = request.form.get('trabalhos_notaveis', '').strip()
        pontos_atencao = request.form.get('pontos_atencao', '').strip()
        valor_medido_str = request.form.get('valor_medido', '0').strip().replace(',', '.')
        avanco_fisico_str = request.form.get('avanco_fisico', '0').strip().replace(',', '.')
        pluviometria = parse_pluviometria(request.form)

        inicio_bl_e    = request.form.get('inicio_bl', '').strip() or None
        fim_bl_e       = request.form.get('fim_bl', '').strip() or None
        data_real_e    = request.form.get('data_real', '').strip() or None
        data_forecast_e= request.form.get('data_forecast', '').strip() or None

        erros = []
        if not contratada:
            erros.append('Contratada é obrigatória.')
        if not semana_referencia:
            erros.append('Semana de referência é obrigatória.')
        if not contrato:
            erros.append('Contrato é obrigatório.')
        if not trabalhos_notaveis:
            erros.append('Trabalhos notáveis são obrigatórios.')

        _bl_ok_e, _bl_err_e = validar_datas_bl(inicio_bl_e, fim_bl_e, data_real_e, data_forecast_e)
        if not _bl_ok_e:
            erros.append(f'Datas BL: {_bl_err_e}')

        try:
            valor_medido = float(valor_medido_str) if valor_medido_str else 0.0
        except:
            valor_medido = 0.0
            erros.append('Valor medido inválido.')

        try:
            avanco_fisico = float(avanco_fisico_str) if avanco_fisico_str else 0.0
        except:
            avanco_fisico = 0.0

        for r in registros:
            if r['id'] != id and r['contratada'] == contratada and r['semana_referencia'] == semana_referencia:
                erros.append(f'Já existe outro registro para "{contratada}" na semana de {format_date_br(semana_referencia)}.')
                break

        efetivo, total_direto, total_indireto = parse_efetivo(request.form)
        equipamentos = parse_equipamentos(request.form)
        acoes_realizadas        = parse_acoes_realizadas(request.form)
        acoes_justificativas    = _parse_json_dict(request.form, 'acoes_justificativas_json')
        acoes_forecast          = _parse_json_dict(request.form, 'acoes_forecast_json')
        equipamentos_realizados = _parse_json_dict(request.form, 'equipamentos_realizados_json')
        histograma_realizados   = _parse_json_dict(request.form, 'histograma_realizados_json')
        avanco_fisico_fc        = _parse_json_dict(request.form, 'avanco_fisico_forecast_json')

        if erros:
            for e in erros:
                flash(e, 'danger')
            return render_template('form.html',
                                   modo='editar',
                                   registro=registro,
                                   contratadas=get_contratadas(),
                                   pluviometria_opcoes=PLUVIOMETRIA_OPCOES,
                                   dias_semana=DIAS_SEMANA,
                                   form_data=request.form,
                                   pluviometria_data=pluviometria,
                                   acoes_realizadas=acoes_realizadas,
                                   acoes_justificativas=acoes_justificativas,
                                   avanco_fisico_forecast=avanco_fisico_fc,
                                   avanco_fisico_fc_manual=avanco_fisico_fc_manual_map(),
                                   equipamentos_realizados=equipamentos_realizados,
                                   histograma_realizados=histograma_realizados,
                                   contratos_cfg_json=json.dumps(load_contratos_config()))

        registro.update({
            'contrato': contrato,
            'contratada': contratada,
            'semana_referencia': semana_referencia,
            'trabalhos_notaveis': trabalhos_notaveis,
            'efetivo': efetivo,
            'total_direto': total_direto,
            'total_indireto': total_indireto,
            'equipamentos': equipamentos,
            'pontos_atencao': pontos_atencao,
            'valor_medido': valor_medido,
            'avanco_fisico': avanco_fisico,
            'avanco_fisico_forecast': avanco_fisico_fc,
            'pluviometria': pluviometria,
            'acoes_realizadas': acoes_realizadas,
            'acoes_justificativas': acoes_justificativas,
            'equipamentos_realizados': equipamentos_realizados,
            'histograma_realizados': histograma_realizados,
            'atualizado_em': datetime.now().isoformat(),
            'alterado_em': datetime.now().isoformat(),
            'alterado_por': current_user_label(),
            'inicio_bl':     inicio_bl_e,
            'fim_bl':        fim_bl_e,
            'data_real':     data_real_e,
            'data_forecast': data_forecast_e,
        })

        save_data(registros)
        recompute_forecast_acoes(contratada, contrato, overrides=acoes_forecast)
        audit_log('editar_registro', id,
                  f'{contratada} / {contrato} / {format_date_br(semana_referencia)}')
        flash('Registro atualizado com sucesso!', 'success')
        return redirect(url_for('index'))

    pluv_existente = registro.get('pluviometria', {})
    if not isinstance(pluv_existente, dict):
        pluv_existente = {}

    return render_template('form.html',
                           modo='editar',
                           registro=registro,
                           contratadas=get_contratadas(),
                           pluviometria_opcoes=PLUVIOMETRIA_OPCOES,
                           dias_semana=DIAS_SEMANA,
                           form_data=registro,
                           pluviometria_data=pluv_existente,
                           acoes_realizadas=registro.get('acoes_realizadas', {}),
                           acoes_justificativas=registro.get('acoes_justificativas', {}),
                           avanco_fisico_forecast=registro.get('avanco_fisico_forecast', {}),
                           avanco_fisico_fc_manual=avanco_fisico_fc_manual_map(exclude_id=id),
                           acoes_acumulado=acoes_acumulado_map(exclude_id=id),
                           acoes_real_sem=acoes_real_semanas_map(exclude_id=id),
                           acoes_fc_manual=acoes_forecast_manual_map(),
                           equipamentos_realizados=registro.get('equipamentos_realizados', {}),
                           histograma_realizados=registro.get('histograma_realizados', {}),
                           contratos_cfg_json=json.dumps(load_contratos_config()))


@app.route('/excluir/<id>', methods=['POST'])
def excluir(id):
    senha_ok = request.form.get('senha') == ADMIN_PASSWORD
    if not can_write() and not senha_ok:
        flash('Sem permissão. Faça login ou informe a senha de administrador.', 'danger')
        return redirect(url_for('index'))

    registros = load_data()
    alvo = next((r for r in registros if r.get('id') == id), None)
    registros = [r for r in registros if r.get('id') != id]
    save_data(registros)

    if alvo:
        ator = current_user_label() if can_write() else 'Administrador (senha)'
        audit_log('excluir_registro', id,
                  f'{alvo.get("contratada","")} / {alvo.get("contrato","")} / '
                  f'{format_date_br(alvo.get("semana_referencia",""))} — por {ator}')
    flash('Registro excluído com sucesso.', 'success')
    return redirect(url_for('index'))


@app.route('/dashboard')
def dashboard():
    registros = load_data()
    todas_contratadas = sorted(set(r.get('contratada', '') for r in registros if r.get('contratada')))

    filtro_contratada = request.args.get('contratada', '')
    filtro_de        = request.args.get('de', '')
    filtro_ate       = request.args.get('ate', '')

    vc = viewer_contratada()
    if vc:
        filtro_contratada = vc
        todas_contratadas = [vc]

    reg = scope_registros(registros[:])  # restringe ao contrato do usuário
    if filtro_contratada:
        reg = [r for r in reg if r.get('contratada') == filtro_contratada]
    if filtro_de:
        reg = [r for r in reg if r.get('semana_referencia', '') >= filtro_de]
    if filtro_ate:
        reg = [r for r in reg if r.get('semana_referencia', '') <= filtro_ate]

    _vazio = dict(kpis=None, contratadas=todas_contratadas,
                  filtro_contratada=filtro_contratada,
                  filtro_de=filtro_de, filtro_ate=filtro_ate,
                  chart_labels='[]', curva_fin_acum='[]', curva_fis_real='[]', curva_fin_base='[]', curva_fis_base='[]',
                  curva_fin_fc='[]', curva_fis_fc='[]', curva_fin_repl='null', curva_fis_repl='null',
                  hist_labels='[]', hist_qtd='[]', hist_colors='[]', hist_list=[],
                  prev_direto=0, prev_indireto=0,
                  acoes_labels='[]', acoes_pct='[]',
                  pluv_labels='[]', pluv_mm='[]')

    if not reg:
        return render_template('dashboard.html', **_vazio)

    reg_ord = sorted(reg, key=lambda r: r.get('semana_referencia', ''))

    # ── KPIs ──
    total_medido    = sum(r.get('valor_medido', 0) for r in reg)
    total_contratos = len(set(r.get('contrato', '') for r in reg if r.get('contrato')))
    ultima_semana   = reg_ord[-1].get('semana_referencia')
    reg_semana      = [r for r in reg if r.get('semana_referencia') == ultima_semana]

    # af_acumulado: último avanço registrado por contrato (valor cumulativo)
    ultimo_por_contrato = {}
    for r in reg_ord:
        chave = r.get('contrato') or r['id']
        ultimo_por_contrato[chave] = r.get('avanco_fisico', 0)
    af_acumulado = (sum(ultimo_por_contrato.values()) / len(ultimo_por_contrato)) if ultimo_por_contrato else 0

    # af_semana: incremento da semana vigente = AF atual - AF semana anterior, por contrato
    af_semana_vals = []
    for r in reg_semana:
        chave = r.get('contrato') or r['id']
        af_atual = r.get('avanco_fisico', 0)
        anterior = next(
            (p.get('avanco_fisico', 0) for p in reversed(reg_ord)
             if (p.get('contrato') or p['id']) == chave
             and p.get('semana_referencia', '') < ultima_semana),
            0
        )
        af_semana_vals.append(af_atual - anterior)
    af_semana = (sum(af_semana_vals) / len(af_semana_vals)) if af_semana_vals else 0

    valor_semana = sum(r.get('valor_medido', 0) for r in reg_semana)

    # ── Dados por semana para Curvas S ──
    semanas_data = {}
    for r in reg_ord:
        sem = r.get('semana_referencia', '')
        if sem not in semanas_data:
            semanas_data[sem] = {'valor': 0, 'af_sum': 0, 'af_n': 0}
        semanas_data[sem]['valor']  += r.get('valor_medido', 0)
        semanas_data[sem]['af_sum'] += r.get('avanco_fisico', 0)
        semanas_data[sem]['af_n']   += 1

    semanas_sorted = sorted(semanas_data.keys())
    chart_labels   = [format_date_br(s) for s in semanas_sorted]

    curva_fin_acum, acum = [], 0
    for s in semanas_sorted:
        acum += semanas_data[s]['valor']
        curva_fin_acum.append(round(acum, 2))

    curva_fis_real = []
    for s in semanas_sorted:
        d   = semanas_data[s]
        avg = d['af_sum'] / d['af_n'] if d['af_n'] else 0
        curva_fis_real.append(round(avg, 2))

    # ── Pluviometria Semanal (mm) ───────────────────────────────────────────
    # Converte os estados qualitativos diários (Tempo Bom, Chuva Improdutiva, …)
    # em mm via PLUVIOMETRIA_MM, soma os 7 dias de cada registro e, quando há mais
    # de um registro na mesma semana, faz a média (a chuva é um fenômeno regional).
    pluv_sem = {}   # semana -> {'soma': mm_total, 'n': qtd_registros}
    for r in reg_ord:
        sem = r.get('semana_referencia', '')
        pluv = r.get('pluviometria') or {}
        if not sem or not isinstance(pluv, dict):
            continue
        mm_semana = sum(PLUVIOMETRIA_MM.get((v or '').strip(), 0) for v in pluv.values())
        acc = pluv_sem.setdefault(sem, {'soma': 0, 'n': 0})
        acc['soma'] += mm_semana
        acc['n']    += 1
    pluv_labels = [format_date_br(s) for s in semanas_sorted]
    pluv_mm     = [round(pluv_sem[s]['soma'] / pluv_sem[s]['n']) if pluv_sem.get(s, {}).get('n') else 0
                   for s in semanas_sorted]

    # ── Histograma consolidado (usa o tipo gravado no registro; reclassifica só se ausente) ──
    histograma = {}
    for r in reg:
        for ef in r.get('efetivo', []):
            funcao = ef.get('funcao', '').strip()
            if not funcao:
                continue
            qtd  = ef.get('quantidade', 0)
            tipo = ef.get('tipo') if ef.get('tipo') in ('direto', 'indireto') else classify_tipo(funcao)
            if funcao not in histograma:
                histograma[funcao] = {'tipo': tipo, 'total': 0}
            histograma[funcao]['total'] += qtd

    hist_list = sorted(histograma.items(), key=lambda x: x[1]['total'], reverse=True)

    TOP = 15
    hist_labels = [k for k, _ in hist_list[:TOP]]
    hist_qtd    = [v['total'] for _, v in hist_list[:TOP]]
    _COLOR_MAP  = {'direto': 'rgba(141,198,63,.85)', 'indireto': 'rgba(0,174,239,.85)', 'classificar': 'rgba(240,165,0,.85)'}
    hist_colors = [_COLOR_MAP.get(v['tipo'], '#ccc') for _, v in hist_list[:TOP]]

    # ── Saldo = Valor do Contrato (ADM) − Total Medido ──
    cfg = load_contratos_config()
    pares_unicos = set(
        contrato_key(r.get('contratada', ''), r.get('contrato', ''))
        for r in reg if r.get('contratada') and r.get('contrato')
    )
    total_valor_contrato = sum(
        valor_efetivo(cfg.get(k, {}))
        for k in pares_unicos
    )
    saldo = total_valor_contrato - total_medido

    # ── Linhas de Base para Curvas S (com Forecast + Replanejamento) ──
    # Estende o eixo do tempo até o fim do baseline/contrato (igual ao Consolidado)
    future_dates = _future_dates(cfg, semanas_sorted, filtro_contratada)
    all_dates    = semanas_sorted + future_dates
    chart_labels = [format_date_br(s) for s in all_dates]
    _nf = len(future_dates)

    # Realizado não existe no futuro
    curva_fin_acum = curva_fin_acum + [None] * _nf
    curva_fis_real = curva_fis_real + [None] * _nf

    curva_fin_base = fin_base_curve(cfg, filtro_contratada, all_dates)

    # Física: o percentual da linha de base já é o avanço ACUMULADO planejado por mês
    # (ex.: 8, 25, 48, 75, 100). Média simples entre contratos e amostra o último mês
    # ≤ a semana (sem cumsum), igual ao Consolidado. Helper reusado p/ baseline e replan.
    def _fis_curve(usar_replan):
        monthly, houve = {}, False
        for _, cdata in cfg.items():
            if filtro_contratada and cdata.get('contratada') != filtro_contratada:
                continue
            base = {it['semana']: float(it.get('percentual', 0) or 0)
                    for it in cdata.get('linha_base_fisica', []) if it.get('semana')}
            if usar_replan:
                base, ch = _mesclar_replan(base, cdata.get('lb_fis_replanejados'))
                houve = houve or ch
            for m, p in base.items():
                monthly.setdefault(m, []).append(float(p or 0))
        cumul = {m: round(sum(vals) / len(vals), 2) for m, vals in monthly.items() if vals}
        months = sorted(cumul)
        curva = [next((cumul[bm] for bm in reversed(months) if bm <= _month_of_week(s)), None)
                 for s in all_dates]
        return curva, houve

    curva_fis_base, _ = _fis_curve(False)

    # Forecast (linha de base deslocada pelo desvio atual)
    curva_fin_fc = _forecast_series(curva_fin_acum, curva_fin_base)
    curva_fis_fc = _forecast_series(curva_fis_real, curva_fis_base, clamp_max=100)

    # Replanejado (só aparece se algum contrato tiver replanejamento)
    curva_fin_base_repl, _hr_fin = fin_base_curve(cfg, filtro_contratada, all_dates, replan=True)
    curva_fin_repl = curva_fin_base_repl if _hr_fin else None
    _fis_repl_curve, _hr_fis = _fis_curve(True)
    curva_fis_repl = _fis_repl_curve if _hr_fis else None

    # ── Histograma Previsto por Função (equipamentos ficam fora do efetivo) ──
    hist_previsto = {}
    for _, cdata in cfg.items():
        if filtro_contratada and cdata.get('contratada') != filtro_contratada:
            continue
        for entry in cdata.get('linha_base_histograma', []):
            funcao = (entry.get('funcao') or '').strip()
            tipo   = entry.get('tipo', 'direto')
            if not funcao or tipo == 'equipamento':
                continue
            total = sum(int(v or 0) for v in entry.get('semanas', {}).values())
            if funcao not in hist_previsto:
                hist_previsto[funcao] = {'tipo': tipo, 'total': 0}
            hist_previsto[funcao]['total'] += total
    prev_list     = sorted(hist_previsto.items(), key=lambda x: x[1]['total'], reverse=True)
    prev_direto   = sum(v['total'] for _, v in prev_list if v['tipo'] == 'direto')
    prev_indireto = sum(v['total'] for _, v in prev_list if v['tipo'] != 'direto')

    # ── Progresso das Ações Notáveis ────────────────────────────────────────
    # Realizado: soma dos valores lançados nos registros da semana (campo "Realizado" do form).
    # Sem nenhum lançamento, usa o previsto acumulado até a última semana como referência.
    real_por_acao = {}
    for r in reg:
        for acao, val in (r.get('acoes_realizadas') or {}).items():
            try:
                real_por_acao[acao] = real_por_acao.get(acao, 0) + float(val or 0)
            except Exception:
                continue
    tem_realizado = bool(real_por_acao)

    last_sem = semanas_sorted[-1] if semanas_sorted else ''
    acoes_prog = {}
    for _, cdata in cfg.items():
        if filtro_contratada and cdata.get('contratada') != filtro_contratada:
            continue
        for entry in cdata.get('linha_base_acoes', []):
            acao = (entry.get('acao') or '').strip()
            vals = entry.get('semanas', {})
            if not acao or not vals:
                continue
            total_plan = sum(float(v or 0) for v in vals.values())
            if total_plan == 0:
                continue
            if tem_realizado:
                done = real_por_acao.get(acao, 0)
            else:
                done = sum(float(v or 0) for k, v in vals.items() if k <= last_sem) if last_sem else 0
            if acao not in acoes_prog:
                acoes_prog[acao] = {'done': 0, 'total': 0}
            acoes_prog[acao]['done']   = done if tem_realizado else acoes_prog[acao]['done'] + done
            acoes_prog[acao]['total'] += total_plan
    acoes_sorted = sorted(acoes_prog.items(), key=lambda x: -(x[1]['done'] / x[1]['total']) if x[1]['total'] else 0)
    acoes_labels = json.dumps([a for a, _ in acoes_sorted])
    acoes_pct    = json.dumps([round(min(100, v['done'] / v['total'] * 100), 1) if v['total'] else 0 for _, v in acoes_sorted])

    kpis = {
        'total_contratos':      total_contratos,
        'total_medido':         total_medido,
        'total_valor_contrato': total_valor_contrato,
        'saldo':                saldo,
        'ultima_semana':        ultima_semana,
        'af_semana':            round(af_semana, 2),
        'af_acumulado':         round(af_acumulado, 2),
        'valor_semana':         valor_semana,
        'valor_acumulado':      total_medido,
    }

    return render_template('dashboard.html',
                           kpis=kpis,
                           contratadas=todas_contratadas,
                           filtro_contratada=filtro_contratada,
                           filtro_de=filtro_de,
                           filtro_ate=filtro_ate,
                           chart_labels=json.dumps(chart_labels),
                           curva_fin_acum=json.dumps(curva_fin_acum),
                           curva_fis_real=json.dumps(curva_fis_real),
                           curva_fin_base=json.dumps(curva_fin_base),
                           curva_fis_base=json.dumps(curva_fis_base),
                           curva_fin_fc=json.dumps(curva_fin_fc),
                           curva_fis_fc=json.dumps(curva_fis_fc),
                           curva_fin_repl=json.dumps(curva_fin_repl),
                           curva_fis_repl=json.dumps(curva_fis_repl),
                           hist_labels=json.dumps(hist_labels),
                           hist_qtd=json.dumps(hist_qtd),
                           hist_colors=json.dumps(hist_colors),
                           hist_list=hist_list,
                           prev_direto=prev_direto,
                           prev_indireto=prev_indireto,
                           acoes_labels=acoes_labels,
                           acoes_pct=acoes_pct,
                           pluv_labels=json.dumps(pluv_labels),
                           pluv_mm=json.dumps(pluv_mm))


def _excel_write_sheet(ws, headers, rows, col_widths, currency_cols=(), percent_cols=()):
    """Escreve cabeçalho estilizado + linhas zebradas + larguras + freeze no padrão do app."""
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    alt_fill    = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
    normal_font = Font(name='Calibri', size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left        = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border      = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
    )

    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, row_data in enumerate(rows, 2):
        fill = alt_fill if i % 2 == 0 else None
        ws.row_dimensions[i].height = 20
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = left
            if fill:
                cell.fill = fill
            if col in currency_cols:
                cell.number_format = '"R$" #,##0.00'
            elif col in percent_cols:
                cell.number_format = '0.00"%"'

    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = 'A2'


@app.route('/export/excel')
def export_excel():
    registros = scope_registros(load_data())
    os.makedirs(EXPORT_DIR, exist_ok=True)

    wb = Workbook()

    # Aba 1: Registros
    ws1 = wb.active
    ws1.title = 'Registros'
    headers1 = ['ID', 'Contrato', 'Contratada', 'Semana Referência', 'Trabalhos Notáveis',
                'Total Direto', 'Total Indireto', 'Pontos de Atenção',
                'Valor Medido da Semana (R$)', 'Avanço Físico (%)', 'Ações Realizadas',
                'Pluv. Segunda', 'Pluv. Terça', 'Pluv. Quarta', 'Pluv. Quinta',
                'Pluv. Sexta', 'Pluv. Sábado', 'Pluv. Domingo',
                'Criado Em', 'Atualizado Em']

    rows1 = []
    for r in registros:
        pluv = r.get('pluviometria', {})
        if not isinstance(pluv, dict):
            pluv = {}
        acoes_str = ' | '.join(
            f'{acao}: {val}%'
            for acao, val in (r.get('acoes_realizadas') or {}).items()
        )
        rows1.append([
            r.get('id', ''),
            r.get('contrato', ''),
            r.get('contratada', ''),
            r.get('semana_referencia', ''),
            r.get('trabalhos_notaveis', ''),
            r.get('total_direto', 0),
            r.get('total_indireto', 0),
            r.get('pontos_atencao', ''),
            r.get('valor_medido', 0),
            r.get('avanco_fisico', 0),
            acoes_str,
            pluv.get('segunda', ''), pluv.get('terca', ''), pluv.get('quarta', ''),
            pluv.get('quinta', ''), pluv.get('sexta', ''), pluv.get('sabado', ''),
            pluv.get('domingo', ''),
            r.get('criado_em', '')[:19].replace('T', ' ') if r.get('criado_em') else '',
            r.get('atualizado_em', '')[:19].replace('T', ' ') if r.get('atualizado_em') else '',
        ])

    _excel_write_sheet(ws1, headers1, rows1,
                       col_widths=[38, 12, 22, 16, 40, 12, 14, 35, 24, 15, 35,
                                   18, 18, 18, 18, 18, 18, 18, 20, 20],
                       currency_cols={9}, percent_cols={10})

    # Aba 2: Histograma Detalhado
    ws2 = wb.create_sheet('Histograma_Detalhado')
    rows2 = [
        [r.get('id', ''), r.get('contratada', ''), r.get('semana_referencia', ''),
         ef.get('funcao', ''), ef.get('quantidade', 0), ef.get('tipo', '')]
        for r in registros for ef in r.get('efetivo', [])
    ]
    _excel_write_sheet(ws2,
                       ['ID Registro', 'Contratada', 'Semana Referência', 'Função', 'Quantidade', 'Tipo'],
                       rows2, col_widths=[38, 22, 16, 28, 12, 12])

    # Aba 3: Equipamentos Detalhados
    ws3 = wb.create_sheet('Equipamentos_Detalhados')
    rows3 = [
        [r.get('id', ''), r.get('contratada', ''), r.get('semana_referencia', ''),
         eq.get('descricao', ''), eq.get('quantidade', 0)]
        for r in registros for eq in r.get('equipamentos', [])
    ]
    _excel_write_sheet(ws3,
                       ['ID Registro', 'Contratada', 'Semana Referência', 'Equipamento', 'Quantidade'],
                       rows3, col_widths=[38, 22, 16, 35, 12])

    filepath = os.path.join(EXPORT_DIR, 'registros_semanais.xlsx')
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name='registros_semanais.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/contratos')
def export_contratos():
    if not _admin_required():
        return redirect(url_for('admin_login'))
    cfg = load_contratos_config()
    os.makedirs(EXPORT_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos'

    rows = []
    for _, data in sorted(cfg.items()):
        status = contract_status(data)
        rows.append([
            data.get('contratada', ''),
            data.get('contrato', ''),
            'Ativo' if status == 'ativo' else 'Encerrado',
            float(data.get('valor_contrato', 0) or 0),
            valor_aditivos(data),
            valor_efetivo(data),
            _week_to_last_day(data.get('data_inicio_contrato', '')),
            _week_to_last_day(data.get('data_fim_contrato', '')),
        ])

    _excel_write_sheet(ws,
                       ['Contratada', 'Contrato', 'Status', 'Valor Base (R$)',
                        'Aditivos (R$)', 'Valor Efetivo (R$)',
                        'Início do Contrato', 'Término do Contrato'],
                       rows, col_widths=[30, 18, 12, 20, 18, 20, 20, 20],
                       currency_cols={4, 5, 6})

    filepath = os.path.join(EXPORT_DIR, 'contratos.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name='contratos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/visualizar/<id>')
def visualizar(id):
    registros = load_data()
    registro = next((r for r in registros if str(r.get('id')) == str(id)), None)
    if not registro:
        flash('Registro não encontrado.', 'danger')
        return redirect(url_for('index'))
    vc = viewer_contratada()
    vk = viewer_contrato()
    if (vc and registro.get('contratada') != vc) or (vk and registro.get('contrato') != vk):
        flash('Você não tem acesso a este registro.', 'danger')
        return redirect(url_for('index'))
    return render_template('visualizar.html', r=registro)


# ── AUTENTICAÇÃO DE USUÁRIOS ────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        user = find_user_by_email(email)
        if user and user.get('ativo', True) and check_password_hash(user.get('senha_hash', ''), senha):
            session.pop('admin_ok', None)
            session['user_id'] = user['id']
            agora = datetime.now().isoformat()
            user['ultimo_login'] = agora
            historico = user.get('historico_logins') or []
            historico.append(agora)
            user['historico_logins'] = historico[-50:]  # mantém os 50 logins mais recentes
            usuarios = load_usuarios()
            for i, u in enumerate(usuarios):
                if u.get('id') == user['id']:
                    usuarios[i] = user
                    break
            save_usuarios(usuarios)
            flash(f'Bem-vindo(a), {user.get("nome") or user.get("email")}!', 'success')
            nxt = request.args.get('next', '')
            if nxt.startswith('/') and not nxt.startswith('//'):
                return redirect(nxt)
            # Após login, vai para a capa (cards de navegação)
            return redirect(url_for('capa'))
        flash('E-mail ou senha inválidos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('admin_ok', None)
    flash('Sessão encerrada.', 'success')
    return redirect(url_for('capa'))


@app.route('/esqueci-senha', methods=['GET', 'POST'])
def esqueci_senha():
    reset_url_dev = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = find_user_by_email(email)
        if user:
            usuarios = load_usuarios()
            for u in usuarios:
                if u.get('id') == user['id']:
                    token = _set_reset_token(u)
                    save_usuarios(usuarios)
                    reset_url = url_for('redefinir_senha', token=token, _external=True)
                    enviado = send_email(u['email'], 'Redefinição de senha — TMS',
                                         _email_reset_html(reset_url))
                    audit_log('reset_senha_solicitado', u['email'])
                    if not enviado:
                        reset_url_dev = reset_url  # fallback: mostra o link na tela
                    break
        # Mensagem genérica (não revela se o e-mail existe)
        if not reset_url_dev:
            flash('Se o e-mail estiver cadastrado, enviamos um link de redefinição.', 'success')
        return render_template('esqueci_senha.html', reset_url_dev=reset_url_dev)
    return render_template('esqueci_senha.html', reset_url_dev=None)


@app.route('/redefinir-senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    usuarios = load_usuarios()
    user = next((u for u in usuarios if u.get('reset_token') == token), None)

    valido = False
    if user:
        exp = user.get('reset_expira')
        try:
            valido = bool(exp) and datetime.fromisoformat(exp) > datetime.now()
        except Exception:
            valido = False

    if not valido:
        return render_template('redefinir_senha.html', valido=False, token=token)

    if request.method == 'POST':
        s1 = request.form.get('senha', '')
        s2 = request.form.get('senha2', '')
        if len(s1) < 6:
            flash('A senha deve ter ao menos 6 caracteres.', 'danger')
        elif s1 != s2:
            flash('As senhas não conferem.', 'danger')
        else:
            user['senha_hash']   = generate_password_hash(s1)
            user['reset_token']  = None
            user['reset_expira'] = None
            save_usuarios(usuarios)
            audit_log('reset_senha_concluido', user['email'])
            flash('Senha redefinida com sucesso. Faça login.', 'success')
            return redirect(url_for('login'))

    return render_template('redefinir_senha.html', valido=True, token=token)


# ── ADMIN ──────────────────────────────────────────────────────────────────

def _admin_required():
    if session.get('admin_ok') is True:
        return True
    u = current_user()
    return bool(u) and u.get('role') in ADMIN_ROLES


def _is_master():
    """True apenas para usuário com role 'master' (exclui admin por senha e rumo)."""
    u = current_user()
    return bool(u) and u.get('role') == 'master'


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    erro = False
    if request.method == 'POST':
        if request.form.get('senha') == ADMIN_PASSWORD:
            session.pop('user_id', None)
            session['admin_ok'] = True
            return redirect(url_for('admin'))
        erro = True
    return render_template('admin_login.html', erro=erro)


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_ok', None)
    return redirect(url_for('index'))


@app.route('/admin/dbinfo')
def admin_dbinfo():
    """Diagnóstico do banco: backend em uso (Postgres vs SQLite) e contagem por tabela."""
    if not _admin_required():
        return redirect(url_for('admin_login'))
    return jsonify(db.backend_info())


@app.route('/admin/dados')
def admin_dados():
    """Tabela completa de registros com filtros e exportação CSV."""
    if not _admin_required():
        return redirect(url_for('admin_login'))

    registros  = load_data()
    contratos  = db.load_contratos()

    # Listas para os selects
    contratadas_list = sorted({r.get('contratada','') for r in registros if r.get('contratada')})
    contratos_list   = sorted({r.get('contrato','')   for r in registros if r.get('contrato')})

    # Filtros
    f_contratada = request.args.get('contratada', '').strip()
    f_contrato   = request.args.get('contrato', '').strip()
    f_de         = request.args.get('de', '').strip()
    f_ate        = request.args.get('ate', '').strip()
    f_criado_por = request.args.get('criado_por', '').strip()

    def _passes(r):
        if f_contratada and r.get('contratada') != f_contratada: return False
        if f_contrato   and r.get('contrato')   != f_contrato:   return False
        semana = r.get('semana_referencia', '')
        if f_de  and semana < f_de:  return False
        if f_ate and semana > f_ate: return False
        if f_criado_por and f_criado_por.lower() not in (r.get('criado_por') or '').lower(): return False
        return True

    filtrados = [r for r in registros if _passes(r)]
    filtrados.sort(key=lambda r: r.get('semana_referencia',''), reverse=True)

    # Contagens para seção Limpar Dados
    counts = db.backend_info().get('counts', {})

    return render_template('admin_dados.html',
        registros=filtrados, total=len(registros),
        contratadas_list=contratadas_list, contratos_list=contratos_list,
        f_contratada=f_contratada, f_contrato=f_contrato,
        f_de=f_de, f_ate=f_ate, f_criado_por=f_criado_por,
        cols=WIPE_COLS, counts=counts)


@app.route('/admin/dados/export-csv')
def admin_dados_export_csv():
    """Exporta registros filtrados como CSV."""
    if not _admin_required():
        return redirect(url_for('admin_login'))

    import csv, io
    registros = load_data()

    f_contratada = request.args.get('contratada', '').strip()
    f_contrato   = request.args.get('contrato', '').strip()
    f_de         = request.args.get('de', '').strip()
    f_ate        = request.args.get('ate', '').strip()
    f_criado_por = request.args.get('criado_por', '').strip()

    def _passes(r):
        if f_contratada and r.get('contratada') != f_contratada: return False
        if f_contrato   and r.get('contrato')   != f_contrato:   return False
        semana = r.get('semana_referencia', '')
        if f_de  and semana < f_de:  return False
        if f_ate and semana > f_ate: return False
        if f_criado_por and f_criado_por.lower() not in (r.get('criado_por') or '').lower(): return False
        return True

    filtrados = [r for r in registros if _passes(r)]
    filtrados.sort(key=lambda r: r.get('semana_referencia',''), reverse=True)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['semana_referencia','contratada','contrato','valor_medido',
                     'avanco_fisico','total_direto','total_indireto',
                     'efetivo_direto','efetivo_indireto',
                     'pluviometria_seg','pluviometria_ter','pluviometria_qua',
                     'pluviometria_qui','pluviometria_sex','pluviometria_sab','pluviometria_dom',
                     'trabalhos_notaveis','pontos_atencao',
                     'criado_em','criado_por','atualizado_em','alterado_por','id'])
    for r in filtrados:
        ef = r.get('efetivo') or []
        ef_dir = sum(e.get('quantidade',0) or 0 for e in ef if e.get('tipo')=='direto')
        ef_ind = sum(e.get('quantidade',0) or 0 for e in ef if e.get('tipo')!='direto')
        pluv   = r.get('pluviometria') or {}
        writer.writerow([
            r.get('semana_referencia',''), r.get('contratada',''), r.get('contrato',''),
            r.get('valor_medido',''), r.get('avanco_fisico',''),
            r.get('total_direto',''), r.get('total_indireto',''),
            ef_dir, ef_ind,
            pluv.get('segunda',''), pluv.get('terca',''), pluv.get('quarta',''),
            pluv.get('quinta',''), pluv.get('sexta',''), pluv.get('sabado',''), pluv.get('domingo',''),
            r.get('trabalhos_notaveis',''), r.get('pontos_atencao',''),
            r.get('criado_em',''), r.get('criado_por',''),
            r.get('atualizado_em',''), r.get('alterado_por',''), r.get('id',''),
        ])

    output.seek(0)
    from flask import Response
    nome = f"dados_{f_contratada or 'todos'}_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(
        '﻿' + output.getvalue(),   # BOM para Excel abrir UTF-8 corretamente
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{nome}"'}
    )


@app.route('/admin/limpar-dados', methods=['GET', 'POST'])
def admin_limpar_dados():
    """Limpa (apaga) coleções da base. Protegido por senha; impede re-seed no próximo boot."""
    if not _admin_required():
        return redirect(url_for('admin_login'))
    if request.method == 'POST':
        if request.form.get('senha') != ADMIN_PASSWORD:
            flash('Senha incorreta — nada foi apagado.', 'danger')
            return redirect(url_for('admin_dados'))
        sel = [c for c, _ in WIPE_COLS if request.form.get('col_' + c)]
        if not sel:
            flash('Selecione ao menos uma coleção.', 'warning')
            return redirect(url_for('admin_dados'))
        removed = db.wipe(sel)
        audit_log('limpar_dados', ', '.join(sel), f'apagados: {removed}')
        total = sum(removed.values())
        flash(f'Base limpa: {total} item(ns) apagado(s) — {removed}. Não será repopulada automaticamente.', 'success')
        return redirect(url_for('admin'))
    # A UI de limpeza vive dentro de /admin/dados; o GET direto só redireciona pra lá.
    return redirect(url_for('admin_dados'))


# ── ADMIN: DADOS TMS ─────────────────────────────────────────────────────────

TMS_CONFIG_FILE = os.path.join('data', 'tms_config.json')

def load_tms_config():
    return db.load_tms()

def save_tms_config(data):
    db.save_tms(data)


# ── Suprimento ───────────────────────────────────────────────────────────────

SUPRIMENTO_FILE = os.path.join('data', 'suprimentos.json')

SUPRIMENTO_STAGES = ['prospeccao', 'qualificacao', 'proposta', 'negociacao', 'contrato']
SUPRIMENTO_STAGE_LABELS = {
    'prospeccao':  'Prospecção',
    'qualificacao': 'Qualificação',
    'proposta':    'Proposta',
    'negociacao':  'Negociação',
    'contrato':    'Contrato',
}
SUPRIMENTO_STAGE_COLORS = {
    'prospeccao':  '#6b7c93',
    'qualificacao': '#2BB2EE',
    'proposta':    '#F0B34D',
    'negociacao':  '#e07b00',
    'contrato':    '#46C26A',
}


def load_suprimentos():
    return db.load_suprimentos()


def save_suprimentos(data):
    db.save_suprimentos(data)


def _parse_valor(raw):
    try:
        return float(str(raw).replace('R$', '').replace('.', '').replace(',', '.').strip())
    except Exception:
        return 0.0


@app.route('/admin/suprimento')
def admin_suprimento():
    # Suprimento agora é uma aba dentro de /admin; mantém a URL antiga funcionando.
    if not _admin_required():
        return redirect(url_for('admin_login'))
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/novo', methods=['POST'])
def admin_suprimento_novo():
    if not _admin_required():
        return redirect(url_for('admin_login'))
    now = datetime.now().isoformat()
    usuario = session.get('usuario_email') or 'Administrador'
    item = {
        'id': str(uuid.uuid4()),
        'descricao':         request.form.get('descricao', '').strip(),
        'contratada':        request.form.get('contratada', '').strip(),
        'objeto':            request.form.get('objeto', '').strip(),
        'valor_estimado':    _parse_valor(request.form.get('valor_estimado', '')),
        'status':            'prospeccao',
        'prioridade':        request.form.get('prioridade', 'media'),
        'data_inicio':       request.form.get('data_inicio', ''),
        'data_prev_contrato': request.form.get('data_prev_contrato', ''),
        'area_contrato':     request.form.get('area_contrato', '').strip(),
        'responsavel':       request.form.get('responsavel', '').strip(),
        'notas':             request.form.get('notas', '').strip(),
        'historico': [{'data': now[:10], 'status': 'prospeccao',
                       'nota': 'Cadastro inicial', 'usuario': usuario}],
        'criado_em':  now,
        'criado_por': usuario,
        'atualizado_em': now,
    }
    data = load_suprimentos()
    data.append(item)
    save_suprimentos(data)
    audit_log('criar_suprimento', item['descricao'], f"Contratada: {item['contratada']}")
    flash(f"Suprimento '{item['descricao']}' cadastrado.", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/editar', methods=['POST'])
def admin_suprimento_editar(sid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_suprimentos()
    item = next((s for s in data if s['id'] == sid), None)
    if not item:
        flash('Suprimento não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    now = datetime.now().isoformat()
    item['descricao']         = request.form.get('descricao', item['descricao']).strip()
    item['contratada']        = request.form.get('contratada', item['contratada']).strip()
    item['objeto']            = request.form.get('objeto', item.get('objeto', '')).strip()
    item['valor_estimado']    = _parse_valor(request.form.get('valor_estimado', ''))
    item['prioridade']        = request.form.get('prioridade', item.get('prioridade', 'media'))
    item['data_inicio']       = request.form.get('data_inicio', item.get('data_inicio', ''))
    item['data_prev_contrato'] = request.form.get('data_prev_contrato', item.get('data_prev_contrato', ''))
    item['area_contrato']     = request.form.get('area_contrato', item.get('area_contrato', '')).strip()
    item['responsavel']       = request.form.get('responsavel', item.get('responsavel', '')).strip()
    item['notas']             = request.form.get('notas', item.get('notas', '')).strip()
    item['atualizado_em']     = now
    save_suprimentos(data)
    audit_log('editar_suprimento', item['descricao'], '')
    flash(f"Suprimento '{item['descricao']}' atualizado.", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/avancar', methods=['POST'])
def admin_suprimento_avancar(sid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_suprimentos()
    item = next((s for s in data if s['id'] == sid), None)
    if not item:
        flash('Suprimento não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    cur = item.get('status', 'prospeccao')
    idx = SUPRIMENTO_STAGES.index(cur) if cur in SUPRIMENTO_STAGES else 0
    if idx < len(SUPRIMENTO_STAGES) - 1:
        novo = SUPRIMENTO_STAGES[idx + 1]
        now  = datetime.now().isoformat()
        usuario = session.get('usuario_email') or 'Administrador'
        nota    = request.form.get('nota', '').strip()
        item['status'] = novo
        item.setdefault('historico', []).append({
            'data': now[:10], 'status': novo,
            'nota': nota or f"Avançou para {SUPRIMENTO_STAGE_LABELS[novo]}",
            'usuario': usuario,
        })
        item['atualizado_em'] = now
        save_suprimentos(data)
        audit_log('avancar_suprimento', item['descricao'], f"→ {novo}")
        flash(f"Movido para «{SUPRIMENTO_STAGE_LABELS[novo]}».", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/recuar', methods=['POST'])
def admin_suprimento_recuar(sid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_suprimentos()
    item = next((s for s in data if s['id'] == sid), None)
    if item:
        cur = item.get('status', 'prospeccao')
        idx = SUPRIMENTO_STAGES.index(cur) if cur in SUPRIMENTO_STAGES else 0
        if idx > 0:
            novo = SUPRIMENTO_STAGES[idx - 1]
            now  = datetime.now().isoformat()
            usuario = session.get('usuario_email') or 'Administrador'
            item['status'] = novo
            item.setdefault('historico', []).append({
                'data': now[:10], 'status': novo,
                'nota': f"Recuou para {SUPRIMENTO_STAGE_LABELS[novo]}",
                'usuario': usuario,
            })
            item['atualizado_em'] = now
            save_suprimentos(data)
            audit_log('recuar_suprimento', item['descricao'], f"→ {novo}")
    flash('Status atualizado.', 'info')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/excluir', methods=['POST'])
def admin_suprimento_excluir(sid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_suprimentos()
    item = next((s for s in data if s['id'] == sid), None)
    if item:
        data = [s for s in data if s['id'] != sid]
        save_suprimentos(data)
        audit_log('excluir_suprimento', item['descricao'], '')
        flash(f"Suprimento '{item['descricao']}' excluído.", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/promover', methods=['POST'])
def admin_suprimento_promover(sid):
    """Promove suprimento a contrato ativo."""
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_suprimentos()
    item = next((s for s in data if s['id'] == sid), None)
    if not item:
        flash('Suprimento não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    contratada  = item.get('contratada', '').strip()
    contrato_id = request.form.get('contrato_id', '').strip()
    if not contratada or not contrato_id:
        flash('Informe o identificador do contrato.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    cfg = load_contratos_config()
    key = contrato_key(contratada, contrato_id)
    if key in cfg:
        flash('Já existe um contrato com essa chave.', 'warning')
        return redirect(url_for('admin') + '?tab=suprimento')
    cfg[key] = {
        'contratada': contratada, 'contrato': contrato_id,
        'valor_contrato': item.get('valor_estimado', 0.0),
        'data_inicio_contrato': '', 'data_fim_contrato': '',
        'status_manual': 'auto',
        'linha_base_financeira': [], 'linha_base_fisica': [],
        'linha_base_histograma': [], 'linha_base_equipamentos': [],
        'linha_base_acoes': [], 'aditivos': [],
    }
    save_contratos_config(cfg)
    data = [s for s in data if s['id'] != sid]
    save_suprimentos(data)
    audit_log('promover_suprimento', item['descricao'], f"→ contrato {contrato_id}")
    flash(f"Contrato '{contrato_id}' criado! Configure datas e linha de base.", 'success')
    return redirect(url_for('admin_contrato', key=key))


# ── Pacotes de Suprimento ──────────────────────────────────────────────────────
# Camada de organização ACIMA dos cards de suprimento. Um pacote vive numa etapa
# do pipeline e agrupa vários processos; ao avançar/recuar, move todos os membros
# de uma vez. Cada suprimento referencia o pacote via campo 'pacote_id' e tem seu
# 'status' mantido em sincronia com a etapa do pacote.

PACOTE_FILE = os.path.join('data', 'pacotes.json')


def load_pacotes():
    return db.load_pacotes()


def save_pacotes(data):
    db.save_pacotes(data)


@app.route('/admin/pacote/novo', methods=['POST'])
def admin_pacote_novo():
    if not _admin_required():
        return redirect(url_for('admin_login'))
    nome = request.form.get('nome', '').strip()
    if not nome:
        flash('Informe um nome para o pacote.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    codigo = request.form.get('codigo', '').strip()
    etapa = request.form.get('status', 'prospeccao')
    if etapa not in SUPRIMENTO_STAGES:
        etapa = 'prospeccao'
    now = datetime.now().isoformat()
    usuario = session.get('usuario_email') or 'Administrador'
    pac = {
        'id': str(uuid.uuid4()),
        'nome': nome,
        'codigo': codigo,
        'status': etapa,
        'historico': [{'data': now[:10], 'status': etapa,
                       'nota': 'Pacote criado', 'usuario': usuario}],
        'criado_em': now, 'criado_por': usuario, 'atualizado_em': now,
    }
    data = load_pacotes()
    data.append(pac)
    save_pacotes(data)
    audit_log('criar_pacote', nome, f"Etapa: {SUPRIMENTO_STAGE_LABELS[etapa]}")
    flash(f"Pacote '{nome}' criado.", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/pacote/<pid>/editar', methods=['POST'])
def admin_pacote_editar(pid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_pacotes()
    pac = next((p for p in data if p['id'] == pid), None)
    if not pac:
        flash('Pacote não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    nome = request.form.get('nome', '').strip()
    if nome:
        pac['nome'] = nome
    pac['codigo'] = request.form.get('codigo', pac.get('codigo', '')).strip()
    pac['atualizado_em'] = datetime.now().isoformat()
    save_pacotes(data)
    audit_log('editar_pacote', pac['nome'], '')
    flash('Pacote atualizado.', 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/pacote/<pid>/excluir', methods=['POST'])
def admin_pacote_excluir(pid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    data = load_pacotes()
    pac = next((p for p in data if p['id'] == pid), None)
    if not pac:
        flash('Pacote não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    # Solta os processos do pacote — mantêm a etapa atual, viram cards avulsos.
    sups = load_suprimentos()
    soltos = 0
    for s in sups:
        if s.get('pacote_id') == pid:
            s.pop('pacote_id', None)
            soltos += 1
    if soltos:
        save_suprimentos(sups)
    data = [p for p in data if p['id'] != pid]
    save_pacotes(data)
    audit_log('excluir_pacote', pac['nome'], f"{soltos} processo(s) liberado(s)")
    flash(f"Pacote '{pac['nome']}' excluído. {soltos} processo(s) mantido(s) soltos.", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


def _mover_pacote(pid, direcao):
    """Move o pacote e todos os seus membros uma etapa (+1 avança, -1 recua)."""
    data = load_pacotes()
    pac = next((p for p in data if p['id'] == pid), None)
    if not pac:
        flash('Pacote não encontrado.', 'danger')
        return
    cur = pac.get('status', 'prospeccao')
    idx = SUPRIMENTO_STAGES.index(cur) if cur in SUPRIMENTO_STAGES else 0
    novo_idx = idx + direcao
    if novo_idx < 0 or novo_idx > len(SUPRIMENTO_STAGES) - 1:
        return
    novo = SUPRIMENTO_STAGES[novo_idx]
    now = datetime.now().isoformat()
    usuario = session.get('usuario_email') or 'Administrador'
    verbo = 'Avançou' if direcao > 0 else 'Recuou'
    pac['status'] = novo
    pac.setdefault('historico', []).append({
        'data': now[:10], 'status': novo,
        'nota': f"{verbo} (pacote) para {SUPRIMENTO_STAGE_LABELS[novo]}", 'usuario': usuario})
    pac['atualizado_em'] = now
    save_pacotes(data)
    sups = load_suprimentos()
    n = 0
    for s in sups:
        if s.get('pacote_id') == pid:
            s['status'] = novo
            s.setdefault('historico', []).append({
                'data': now[:10], 'status': novo,
                'nota': f"{verbo} com o pacote «{pac['nome']}»", 'usuario': usuario})
            s['atualizado_em'] = now
            n += 1
    if n:
        save_suprimentos(sups)
    audit_log('mover_pacote', pac['nome'], f"→ {novo} ({n} processo(s))")
    flash(f"Pacote «{pac['nome']}» e {n} processo(s) movidos para «{SUPRIMENTO_STAGE_LABELS[novo]}».", 'success')


@app.route('/admin/pacote/<pid>/avancar', methods=['POST'])
def admin_pacote_avancar(pid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    _mover_pacote(pid, +1)
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/pacote/<pid>/recuar', methods=['POST'])
def admin_pacote_recuar(pid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    _mover_pacote(pid, -1)
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/pacote/<pid>/add', methods=['POST'])
def admin_pacote_add(pid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    pacs = load_pacotes()
    pac = next((p for p in pacs if p['id'] == pid), None)
    if not pac:
        flash('Pacote não encontrado.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    sid = request.form.get('sup_id', '').strip()
    sups = load_suprimentos()
    item = next((s for s in sups if s['id'] == sid), None)
    if not item:
        flash('Selecione um processo válido.', 'danger')
        return redirect(url_for('admin') + '?tab=suprimento')
    now = datetime.now().isoformat()
    usuario = session.get('usuario_email') or 'Administrador'
    destino = pac.get('status', 'prospeccao')
    item['pacote_id'] = pid
    item['status'] = destino
    item.setdefault('historico', []).append({
        'data': now[:10], 'status': destino,
        'nota': f"Incluído no pacote «{pac['nome']}»", 'usuario': usuario})
    item['atualizado_em'] = now
    save_suprimentos(sups)
    audit_log('add_pacote', pac['nome'], item['descricao'])
    flash(f"«{item['descricao']}» incluído no pacote «{pac['nome']}».", 'success')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/suprimento/<sid>/sair-pacote', methods=['POST'])
def admin_suprimento_sair_pacote(sid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    sups = load_suprimentos()
    item = next((s for s in sups if s['id'] == sid), None)
    if item and item.get('pacote_id'):
        item.pop('pacote_id', None)
        item['atualizado_em'] = datetime.now().isoformat()
        save_suprimentos(sups)
        audit_log('remover_de_pacote', item['descricao'], '')
        flash(f"«{item['descricao']}» removido do pacote.", 'info')
    return redirect(url_for('admin') + '?tab=suprimento')


@app.route('/admin/dados-tms', methods=['GET', 'POST'])
@app.route('/admin/tms', methods=['GET', 'POST'])  # alias legado
def admin_tms():
    if not _is_master():
        flash('Acesso restrito ao usuário Master.', 'danger')
        return redirect(url_for('dashboard'))

    cfg = load_tms_config()

    if request.method == 'POST':
        milestones_raw = request.form.get('milestones_json', '[]')
        try:
            milestones = json.loads(milestones_raw)
        except Exception:
            milestones = []
        cfg.update({
            'nome_programa':  request.form.get('nome_programa', '').strip(),
            'sigla':          request.form.get('sigla', '').strip(),
            'cliente':        request.form.get('cliente', '').strip(),
            'objeto':         request.form.get('objeto', '').strip(),
            'responsavel':    request.form.get('responsavel', '').strip(),
            'data_inicio':    request.form.get('data_inicio', '').strip(),
            'data_fim':       request.form.get('data_fim', '').strip(),
            'valor_programa': request.form.get('valor_programa', '').strip(),
            'local':          request.form.get('local', '').strip(),
            'milestones':     milestones,
            'atualizado_em':  datetime.now().isoformat(),
            'atualizado_por': current_user_label(),
        })
        # Linhas de Base mensais (mesma lógica de Configurar Contrato)
        _parse_baseline_mensal(request.form, cfg)
        save_tms_config(cfg)
        audit_log('editar_tms', 'tms_config', 'Dados TMS atualizados')
        flash('Dados TMS salvos com sucesso!', 'success')
        return redirect(url_for('admin_tms'))

    return render_template('admin_tms.html', cfg=cfg)


def _parse_int(valor):
    """Converte texto de formulário para inteiro >= 0 (0 se vazio/ inválido)."""
    try:
        n = int(float(str(valor).replace(',', '.').strip()))
        return max(0, n)
    except (TypeError, ValueError):
        return 0


@app.route('/admin/dados-tms/saude', methods=['POST'])
def admin_tms_saude():
    """Registra (upsert por mês) um lançamento mensal de Saúde e Segurança."""
    if not _is_master():
        flash('Acesso restrito ao usuário Master.', 'danger')
        return redirect(url_for('dashboard'))

    mes_ref = request.form.get('mes_ref', '').strip()  # 'YYYY-MM'
    if not mes_ref:
        flash('Informe o mês/ano de referência.', 'danger')
        return redirect(url_for('admin_tms'))

    registro = {
        'mes_ref':            mes_ref,
        'saf':                _parse_int(request.form.get('saf')),
        'caf':                _parse_int(request.form.get('caf')),
        'primeiros_socorros': _parse_int(request.form.get('primeiros_socorros')),
        'registrado_em':      datetime.now().isoformat(),
        'registrado_por':     current_user_label(),
    }

    cfg = load_tms_config()
    lista = [x for x in cfg.get('saude_seguranca', []) if isinstance(x, dict)]
    # upsert: substitui o mês existente, senão adiciona
    lista = [x for x in lista if x.get('mes_ref') != mes_ref]
    lista.append(registro)
    lista.sort(key=lambda x: x.get('mes_ref', ''), reverse=True)
    cfg['saude_seguranca'] = lista
    save_tms_config(cfg)
    audit_log('editar_tms', 'saude_seguranca', f'Saúde e Segurança — {mes_ref}')
    flash(f'Dados de Saúde e Segurança de {mes_ref} salvos com sucesso!', 'success')
    return redirect(url_for('admin_tms') + '#saude-seguranca')


@app.route('/admin/dados-tms/saude/<mes_ref>/excluir', methods=['POST'])
def admin_tms_saude_excluir(mes_ref):
    if not _is_master():
        flash('Acesso restrito ao usuário Master.', 'danger')
        return redirect(url_for('dashboard'))
    cfg = load_tms_config()
    lista = [x for x in cfg.get('saude_seguranca', []) if isinstance(x, dict)]
    cfg['saude_seguranca'] = [x for x in lista if x.get('mes_ref') != mes_ref]
    save_tms_config(cfg)
    audit_log('editar_tms', 'saude_seguranca', f'Excluído Saúde e Segurança — {mes_ref}')
    flash(f'Registro de {mes_ref} removido.', 'info')
    return redirect(url_for('admin_tms') + '#saude-seguranca')


# ── ADMIN: GESTÃO DE USUÁRIOS ───────────────────────────────────────────────

@app.route('/admin/usuarios')
def admin_usuarios():
    if not _admin_required():
        return redirect(url_for('admin_login'))
    usuarios = sorted(load_usuarios(), key=lambda u: (u.get('contratada') or '', u.get('email', '')))
    cred = session.pop('_nova_credencial', None)  # credenciais geradas (fallback sem SMTP)
    cfg = load_contratos_config()
    contratos_por_contratada = {}
    for cdata in cfg.values():
        c, k = cdata.get('contratada'), cdata.get('contrato')
        if c and k:
            contratos_por_contratada.setdefault(c, [])
            if k not in contratos_por_contratada[c]:
                contratos_por_contratada[c].append(k)

    # Log de atividade — somente criação, edição e exclusão (mais recentes primeiro).
    # Visível apenas para o usuário master.
    cu = current_user()
    auditoria = []
    if cu and cu.get('role') == 'master':
        auditoria = [
            a for a in load_auditoria()
            if str(a.get('acao', '')).startswith(('criar_', 'editar_', 'excluir_'))
        ]
        auditoria = sorted(auditoria, key=lambda a: a.get('data_hora', ''), reverse=True)[:200]
        for a in auditoria:
            a['descricao'] = _audit_descricao(a)

    return render_template('usuarios.html',
                           usuarios=usuarios,
                           contratadas=get_contratadas(),
                           contratos_por_contratada=contratos_por_contratada,
                           smtp_ok=_smtp_configured(),
                           auditoria=auditoria,
                           cred=cred)


@app.route('/admin/usuarios/novo', methods=['POST'])
def admin_usuarios_novo():
    if not _admin_required():
        return redirect(url_for('admin_login'))

    email      = request.form.get('email', '').strip().lower()
    tipo       = request.form.get('tipo', 'contratada')   # contratada | contratada_rw | staff | master | rumo
    contratada = request.form.get('contratada', '').strip()
    contrato   = request.form.get('contrato', '').strip()

    role = {'master': 'master', 'rumo': 'rumo', 'staff': 'staff',
            'contratada_rw': 'contratada_rw'}.get(tipo, 'contratada')
    is_scoped = role in SCOPED_ROLES

    if not email or '@' not in email:
        flash('Informe um e-mail válido.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if role == 'master' and _actor_is_rumo():
        flash('Você não tem permissão para criar usuários master.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if is_scoped and (not contratada or not contrato):
        flash('Selecione a contratada e o contrato vinculados ao usuário.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if find_user_by_email(email):
        flash(f'Já existe um usuário com o e-mail "{email}".', 'warning')
        return redirect(url_for('admin_usuarios'))

    senha = _gen_password()
    usuarios = load_usuarios()
    novo = {
        'id':           str(uuid.uuid4()),
        'email':        email,
        'nome':         '',
        'role':         role,
        'contratada':   contratada if is_scoped else None,
        'contrato':     contrato if is_scoped else None,
        'senha_hash':   generate_password_hash(senha),
        'ativo':        True,
        'reset_token':  None,
        'reset_expira': None,
        'criado_em':    datetime.now().isoformat(),
        'criado_por':   current_user_label(),
        'ultimo_login': None,
        'historico_logins': [],
    }
    token = _set_reset_token(novo)
    usuarios.append(novo)
    save_usuarios(usuarios)
    audit_log('criar_usuario', email,
              f'role={novo["role"]} contratada={contratada or "-"} contrato={contrato or "-"}')

    login_url = url_for('login', _external=True)
    reset_url = url_for('redefinir_senha', token=token, _external=True)
    enviado = send_email(email, 'Seu acesso ao Monitoramento de Obras — TMS',
                         _email_boas_vindas_html('', email, senha, login_url, reset_url))

    if enviado:
        flash(f'Usuário "{email}" criado e e-mail enviado com a senha e o link de redefinição.', 'success')
    else:
        # Fallback: SMTP não configurado — mostra as credenciais uma vez na tela
        session['_nova_credencial'] = {
            'email': email, 'senha': senha,
            'login_url': login_url, 'reset_url': reset_url,
        }
        flash(f'Usuário "{email}" criado. E-mail não configurado — exibindo credenciais abaixo (copie agora).', 'warning')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<uid>/reenviar', methods=['POST'])
def admin_usuarios_reenviar(uid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    usuarios = load_usuarios()
    user = next((u for u in usuarios if u.get('id') == uid), None)
    if not user:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if not can_manage_user(user):
        flash('Você não tem permissão para alterar um usuário master.', 'danger')
        return redirect(url_for('admin_usuarios'))
    token = _set_reset_token(user)
    save_usuarios(usuarios)
    audit_log('reenviar_acesso', user['email'])
    reset_url = url_for('redefinir_senha', token=token, _external=True)
    enviado = send_email(user['email'], 'Redefinição de senha — TMS', _email_reset_html(reset_url))
    if enviado:
        flash(f'Link de redefinição enviado para "{user["email"]}".', 'success')
    else:
        session['_nova_credencial'] = {
            'email': user['email'], 'senha': None,
            'login_url': url_for('login', _external=True), 'reset_url': reset_url,
        }
        flash('E-mail não configurado — exibindo o link de redefinição abaixo.', 'warning')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<uid>/toggle', methods=['POST'])
def admin_usuarios_toggle(uid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    usuarios = load_usuarios()
    user = next((u for u in usuarios if u.get('id') == uid), None)
    if user and not can_manage_user(user):
        flash('Você não tem permissão para alterar um usuário master.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if user:
        user['ativo'] = not user.get('ativo', True)
        save_usuarios(usuarios)
        audit_log('ativar_usuario' if user['ativo'] else 'desativar_usuario', user['email'])
        flash(f'Usuário "{user["email"]}" {"ativado" if user["ativo"] else "desativado"}.', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<uid>/editar', methods=['POST'])
def admin_usuarios_editar(uid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    usuarios = load_usuarios()
    user = next((u for u in usuarios if u.get('id') == uid), None)
    if not user:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if not can_manage_user(user):
        flash('Você não tem permissão para alterar um usuário master.', 'danger')
        return redirect(url_for('admin_usuarios'))

    tipo       = request.form.get('tipo', user.get('role', 'contratada'))
    contratada = request.form.get('contratada', '').strip()
    contrato   = request.form.get('contrato', '').strip()

    role = {'master': 'master', 'rumo': 'rumo', 'staff': 'staff',
            'contratada_rw': 'contratada_rw'}.get(tipo, 'contratada')
    is_scoped = role in SCOPED_ROLES

    if role == 'master' and _actor_is_rumo():
        flash('Você não tem permissão para promover usuários a master.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if is_scoped and (not contratada or not contrato):
        flash('Selecione a contratada e o contrato vinculados ao usuário.', 'danger')
        return redirect(url_for('admin_usuarios'))

    antes = f'{user.get("role")}/{user.get("contratada") or "-"}/{user.get("contrato") or "-"}'
    user['role']       = role
    user['contratada'] = contratada if is_scoped else None
    user['contrato']   = contrato if is_scoped else None
    save_usuarios(usuarios)
    depois = f'{role}/{contratada or "-"}/{contrato or "-"}'
    audit_log('editar_usuario', user['email'], f'{antes} -> {depois}')
    flash(f'Permissões de "{user["email"]}" atualizadas.', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<uid>/excluir', methods=['POST'])
def admin_usuarios_excluir(uid):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    usuarios = load_usuarios()
    user = next((u for u in usuarios if u.get('id') == uid), None)
    if user and not can_manage_user(user):
        flash('Você não tem permissão para excluir um usuário master.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if user:
        usuarios = [u for u in usuarios if u.get('id') != uid]
        save_usuarios(usuarios)
        audit_log('excluir_usuario', user['email'])
        flash(f'Usuário "{user["email"]}" excluído.', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin')
def admin():
    if not _admin_required():
        return redirect(url_for('admin_login'))
    registros = load_data()
    cfg = load_contratos_config()
    contratos = {}
    for r in registros:
        k = contrato_key(r.get('contratada', ''), r.get('contrato', ''))
        if k not in contratos:
            contratos[k] = {
                'key': k,
                'contratada': r.get('contratada', ''),
                'contrato': r.get('contrato', ''),
                'config': cfg.get(k, {}),
            }
    for k, c in cfg.items():
        if k not in contratos:
            contratos[k] = {
                'key': k,
                'contratada': c.get('contratada', ''),
                'contrato': c.get('contrato', ''),
                'config': c,
            }
    suprimentos = load_suprimentos()
    pacotes     = load_pacotes()
    pac_ids     = {p['id'] for p in pacotes}

    # Membros de cada pacote (suprimentos que apontam para ele)
    membros_idx = {}
    for s in suprimentos:
        pidv = s.get('pacote_id')
        if pidv:
            membros_idx.setdefault(pidv, []).append(s)
    for p in pacotes:
        p['membros'] = membros_idx.get(p['id'], [])
        p['valor_total'] = sum((m.get('valor_estimado') or 0) for m in p['membros'])

    # Pacotes agrupados por etapa (para o kanban)
    pacotes_por_status = {s: [] for s in SUPRIMENTO_STAGES}
    for p in pacotes:
        pacotes_por_status.setdefault(p.get('status', 'prospeccao'), []).append(p)

    # por_status = TODOS os suprimentos por etapa (funil/contagens)
    # standalone_por_status = apenas os soltos (sem pacote) — cards avulsos do kanban
    por_status            = {s: [] for s in SUPRIMENTO_STAGES}
    standalone_por_status = {s: [] for s in SUPRIMENTO_STAGES}
    for s in suprimentos:
        st = s.get('status', 'prospeccao')
        por_status.setdefault(st, []).append(s)
        if not s.get('pacote_id') or s.get('pacote_id') not in pac_ids:
            standalone_por_status.setdefault(st, []).append(s)

    # Processos soltos — disponíveis para incluir num pacote
    suprimentos_soltos = [s for s in suprimentos
                          if not s.get('pacote_id') or s.get('pacote_id') not in pac_ids]

    return render_template('admin.html',
                           contratos=sorted(contratos.values(), key=lambda x: x['contratada']),
                           suprimentos=suprimentos,
                           por_status=por_status,
                           standalone_por_status=standalone_por_status,
                           pacotes=pacotes,
                           pacotes_por_status=pacotes_por_status,
                           suprimentos_soltos=suprimentos_soltos,
                           stages=SUPRIMENTO_STAGES,
                           stage_labels=SUPRIMENTO_STAGE_LABELS,
                           stage_colors=SUPRIMENTO_STAGE_COLORS)


@app.route('/admin/novo_contrato', methods=['POST'])
def admin_novo_contrato():
    if not _admin_required():
        return redirect(url_for('admin_login'))

    contratada = request.form.get('contratada', '').strip()
    contrato_id = request.form.get('contrato', '').strip()

    if not contratada or not contrato_id:
        flash('Contratada e Contrato são obrigatórios.', 'danger')
        return redirect(url_for('admin'))

    key = contrato_key(contratada, contrato_id)
    cfg = load_contratos_config()

    if key in cfg:
        flash(f'Contrato "{contrato_id}" para "{contratada}" já existe.', 'warning')
        return redirect(url_for('admin_contrato', key=key))

    cfg[key] = {'contratada': contratada, 'contrato': contrato_id}
    save_contratos_config(cfg)
    audit_log('criar_contrato', key, f'{contratada} / {contrato_id}')
    flash(f'Contrato "{contrato_id}" criado com sucesso. Configure os dados abaixo.', 'success')
    return redirect(url_for('admin_contrato', key=key))


@app.route('/admin/contrato/<path:key>/excluir', methods=['POST'])
def admin_contrato_excluir(key):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    if not _is_master():
        flash('Apenas o usuário master pode excluir contratos.', 'danger')
        return redirect(url_for('admin'))

    cfg = load_contratos_config()
    data = cfg.get(key, {})
    contratada  = data.get('contratada') or key.split('||')[0]
    contrato_id = data.get('contrato') or (key.split('||')[1] if '||' in key else '')

    # Bloqueia exclusão se houver registros vinculados (evita órfãos)
    registros = load_data()
    n_reg = sum(1 for r in registros
                if contrato_key(r.get('contratada', ''), r.get('contrato', '')) == key)
    if n_reg:
        flash(f'Não é possível excluir: há {n_reg} registro(s) vinculado(s) a este contrato. '
              f'Exclua os registros antes.', 'danger')
        return redirect(url_for('admin'))

    if key in cfg:
        del cfg[key]
        save_contratos_config(cfg)
    audit_log('excluir_contrato', key, f'{contratada} / {contrato_id}')
    flash(f'Contrato "{contrato_id}" de "{contratada}" excluído.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/contrato/<path:key>', methods=['GET', 'POST'])
def admin_contrato(key):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    cfg = load_contratos_config()

    if request.method == 'POST':
        data = cfg.get(key, {})
        data['contratada']    = request.form.get('contratada', '')
        data['contrato']      = request.form.get('contrato', '')
        raw_val               = request.form.get('valor_contrato', '').replace('.', '').replace(',', '.')
        data['valor_contrato'] = float(raw_val) if raw_val else 0.0

        # ── Linhas de Base mensais (Financeira + Física) ──────────────────────
        _parse_baseline_mensal(request.form, data)

        data['data_inicio_contrato'] = request.form.get('data_inicio_contrato', '')
        data['data_fim_contrato']    = request.form.get('data_fim_contrato', '')
        data['status_manual']        = request.form.get('status_manual', 'auto')
        data['area_contrato']        = request.form.get('area_contrato', '')

        # ── Classificação financeira (CAPEX/OPEX, centro de custo, prazos) ────
        data['classificacao'] = request.form.get('classificacao', 'CAPEX')
        data['centro_custo']  = request.form.get('centro_custo', '')
        try:
            data['prazo_pagamento_dias'] = max(0, int(request.form.get('prazo_pagamento_dias', '30') or 30))
        except (TypeError, ValueError):
            data['prazo_pagamento_dias'] = 30
        data['retencao_pct'] = parse_brl(request.form.get('retencao_pct', '0'))
        data['impostos_pct'] = parse_brl(request.form.get('impostos_pct', '0'))

        # ── Aditivos (valor / prazo) ── bloco removido da UI; só atualiza se enviado,
        # preservando aditivos já existentes no contrato.
        if 'adit_tipo[]' in request.form:
            adit_tipo  = request.form.getlist('adit_tipo[]')
            adit_valor = request.form.getlist('adit_valor[]')
            adit_prazo = request.form.getlist('adit_prazo[]')
            adit_data  = request.form.getlist('adit_data[]')
            adit_desc  = request.form.getlist('adit_desc[]')
            aditivos = []
            for t, v, p, d, ds in zip(adit_tipo, adit_valor, adit_prazo, adit_data, adit_desc):
                if t not in ('valor', 'prazo'):
                    continue
                valor = float((v or '0').replace('.', '').replace(',', '.')) if t == 'valor' else 0.0
                prazo = p if t == 'prazo' else ''
                # ignora linha totalmente vazia
                if t == 'valor' and not valor and not ds.strip():
                    continue
                if t == 'prazo' and not prazo and not ds.strip():
                    continue
                aditivos.append({'tipo': t, 'valor': valor, 'prazo': prazo,
                                 'data': d, 'descricao': ds.strip()})
            data['aditivos'] = aditivos

        hist_json = request.form.get('hist_json', '[]')
        try:
            hist_rows = json.loads(hist_json)
        except (json.JSONDecodeError, ValueError):
            hist_rows = []

        # Linhas com categoria "equipamento" vão para linha_base_equipamentos
        # (formato {equipamento, semanas}, lido pelo bloco Equipamentos do form).
        data['linha_base_histograma'] = [r for r in hist_rows if r.get('tipo') != 'equipamento']
        data['linha_base_equipamentos'] = [
            {'equipamento': r.get('funcao', ''), 'semanas': r.get('semanas', {})}
            for r in hist_rows if r.get('tipo') == 'equipamento' and r.get('funcao')
        ]

        acoes_json = request.form.get('acoes_json', '[]')
        try:
            data['linha_base_acoes'] = json.loads(acoes_json)
        except (json.JSONDecodeError, ValueError):
            data['linha_base_acoes'] = []

        # Semanas extras adicionadas manualmente no bloco de Ações Notáveis
        try:
            data['acoes_semanas_extra'] = json.loads(request.form.get('acoes_extra_json', '[]'))
        except (json.JSONDecodeError, ValueError):
            data['acoes_semanas_extra'] = []
        # Nº de linhas de "Replanejado" adicionadas pelo botão (0–3)
        try:
            data['acoes_replanejado_n'] = max(0, min(3, int(request.form.get('acoes_repl_n', '0') or 0)))
        except (TypeError, ValueError):
            data['acoes_replanejado_n'] = 0

        cfg[key] = data
        save_contratos_config(cfg)
        audit_log('editar_contrato', key, f'{data.get("contratada","")} / {data.get("contrato","")}')
        flash('Configuração salva com sucesso.', 'success')
        return redirect(url_for('admin'))

    data = cfg.get(key, {})
    if not data.get('contratada'):
        parts = key.split('||', 1)
        data['contratada'] = parts[0]
        data['contrato']   = parts[1] if len(parts) > 1 else ''

    # Mescla equipamentos de volta na tabela do histograma para edição
    data_view = dict(data)
    data_view['linha_base_histograma'] = list(data.get('linha_base_histograma', [])) + [
        {'funcao': e.get('equipamento', ''), 'tipo': 'equipamento', 'semanas': e.get('semanas', {})}
        for e in data.get('linha_base_equipamentos', [])
    ]

    # Real por ação/semana, por mês (financeiro e físico) — lidos dos registros deste contrato.
    acoes_real = {}
    fin_real   = {}   # {YYYY-MM: soma de valor_medido}
    fis_real_list = {}  # {YYYY-MM: [(semana_ref, avanco_fisico)]}
    for r in load_data():
        if r.get('contratada') != data.get('contratada') or r.get('contrato') != data.get('contrato'):
            continue
        semana = r.get('semana_referencia', '')
        wk     = format_date_to_week(semana)
        month  = semana[:7] if len(semana) >= 7 else None
        # Ações notáveis
        if wk:
            for acao, val in (r.get('acoes_realizadas') or {}).items():
                try:
                    acoes_real.setdefault(acao, {})[wk] = float(val)
                except (TypeError, ValueError):
                    pass
        # Financeiro: soma valor_medido por mês
        if month:
            try:
                fin_real[month] = fin_real.get(month, 0) + float(r.get('valor_medido') or 0)
            except (TypeError, ValueError):
                pass
            # Físico: mantém último avanco_fisico por mês
            af = r.get('avanco_fisico')
            if af is not None:
                try:
                    fis_real_list.setdefault(month, []).append((semana, float(af)))
                except (TypeError, ValueError):
                    pass
    fis_real = {m: sorted(v, key=lambda x: x[0])[-1][1] for m, v in fis_real_list.items()}

    return render_template('admin_contrato.html', key=key, contrato=data_view,
                           funcoes_mao_obra=get_funcoes_list(),
                           acoes_real=acoes_real,
                           fin_real=fin_real,
                           fis_real=fis_real,
                           centros_custo=[c for c in load_centros_custo() if c.get('ativo', True)])


# ── Helpers: gera listas de semanas/meses a partir de YYYY-Wnn ──────────────

def _gen_weeks_py(start_yw, end_yw):
    """Retorna lista de 'YYYY-Wnn' de start_yw até end_yw inclusive."""
    import datetime as _dt
    from datetime import timedelta
    weeks = []
    try:
        cur  = _dt.datetime.strptime(get_monday(start_yw), '%Y-%m-%d').date()
        last = _dt.datetime.strptime(get_monday(end_yw),   '%Y-%m-%d').date()
        while cur <= last:
            weeks.append(format_date_to_week(cur.strftime('%Y-%m-%d')))
            cur += timedelta(days=7)
    except Exception:
        pass
    return weeks


def _gen_months_py(start_yw, end_yw):
    """Retorna lista de 'YYYY-MM' cobrindo os meses de start_yw até end_yw inclusive."""
    import datetime as _dt
    months = []
    try:
        sd = _dt.datetime.strptime(get_monday(start_yw), '%Y-%m-%d').date()
        ed = _dt.datetime.strptime(get_monday(end_yw),   '%Y-%m-%d').date()
        y, m = sd.year, sd.month
        while (y, m) <= (ed.year, ed.month):
            months.append(f'{y}-{m:02d}')
            m += 1
            if m > 12:
                m = 1; y += 1
    except Exception:
        pass
    return months


# ── Download: modelo Excel de planejamento ────────────────────────────────────

@app.route('/admin/contrato/<path:key>/template')
def admin_contrato_template(key):
    if not _admin_required():
        return redirect(url_for('admin_login'))
    cfg  = load_contratos_config()
    data = cfg.get(key, {})
    inicio = data.get('data_inicio_contrato', '')
    fim    = data.get('data_fim_contrato',    '')
    weeks  = _gen_weeks_py(inicio, fim)
    months = _gen_months_py(inicio, fim)

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    NAVY     = PatternFill('solid', fgColor='001f4d')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
    LBL_FONT = Font(bold=True, size=9, name='Calibri')
    DAT_FONT = Font(size=9, name='Calibri')
    CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _head(ws, labels):
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row=1, column=ci, value=lbl)
            c.fill = NAVY; c.font = HDR_FONT; c.alignment = CENTER

    def _row(ws, ri, label, vals):
        ws.cell(row=ri, column=1, value=label).font = LBL_FONT
        for ci, v in enumerate(vals, 2):
            if v is not None:
                ws.cell(row=ri, column=ci, value=v).font = DAT_FONT

    def _widths(ws, fixed, n_period, pw=10):
        for col_ltr, w in fixed.items():
            ws.column_dimensions[col_ltr].width = w
        for ci in range(len(fixed) + 1, len(fixed) + n_period + 1):
            ws.column_dimensions[get_column_letter(ci)].width = pw

    # ── Instruções ──────────────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = 'Instrucoes'
    ws0.column_dimensions['A'].width = 72
    linhas = [
        ('MODELO DE IMPORTAÇÃO — PLANEJAMENTO', Font(bold=True, size=13, color='001f4d', name='Calibri')),
        ('', None),
        ('Como usar:', Font(bold=True, size=10, name='Calibri')),
        ('1. Preencha apenas as linhas "Planejado" (obrigatório). Forecast e Replanejado são opcionais.', DAT_FONT),
        ('2. Financeira: valores em R$ (número inteiro, sem ponto de milhar — ex: 60000).', DAT_FONT),
        ('3. Fisica: percentual de 0 a 100 (ex: 25.5).', DAT_FONT),
        ('4. Histograma: quantidade de pessoas por semana (número inteiro).', DAT_FONT),
        ('5. Acoes: percentual de execução previsto por semana (ex: 15.0).', DAT_FONT),
        ('6. Não altere os cabeçalhos YYYY-MM (Financeira/Física) ou YYYY-Wnn (Histograma/Ações).', DAT_FONT),
        ('', None),
        ('Categorias do Histograma: direto | indireto | equipamento', Font(bold=True, size=9, name='Calibri')),
        ('Unidades das Ações: m² | m³ | m | Unid | (deixe em branco para sem unidade)', Font(bold=True, size=9, name='Calibri')),
    ]
    for ri, (txt, fnt) in enumerate(linhas, 1):
        c = ws0.cell(row=ri, column=1, value=txt)
        if fnt: c.font = fnt

    # ── Financeira ───────────────────────────────────────────────────────────
    ws1 = wb.create_sheet('Financeira')
    _head(ws1, ['Série'] + months)
    _widths(ws1, {'A': 16}, len(months))
    fp = data.get('lb_fin_planejado', {}) or {it['semana']: it['valor'] for it in data.get('linha_base_financeira', [])}
    ff = data.get('lb_fin_forecast',  {})
    fr = data.get('lb_fin_replanejados', [])
    _row(ws1, 2, 'Planejado', [fp.get(m) for m in months])
    _row(ws1, 3, 'Forecast',  [ff.get(m) for m in months])
    for i in range(3):
        rp = fr[i] if i < len(fr) else {}
        _row(ws1, 4 + i, f'Replanejado {i + 1}', [rp.get(m) for m in months])

    # ── Fisica ───────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Fisica')
    _head(ws2, ['Série'] + months)
    _widths(ws2, {'A': 16}, len(months))
    gp = data.get('lb_fis_planejado', {}) or {it['semana']: it['percentual'] for it in data.get('linha_base_fisica', [])}
    gf = data.get('lb_fis_forecast',  {})
    gr = data.get('lb_fis_replanejados', [])
    _row(ws2, 2, 'Planejado', [gp.get(m) for m in months])
    _row(ws2, 3, 'Forecast',  [gf.get(m) for m in months])
    for i in range(3):
        rp = gr[i] if i < len(gr) else {}
        _row(ws2, 4 + i, f'Replanejado {i + 1}', [rp.get(m) for m in months])

    # ── Histograma ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Histograma')
    _head(ws3, ['Função', 'Categoria'] + weeks)
    _widths(ws3, {'A': 24, 'B': 14}, len(weeks), pw=9)
    hist_all = list(data.get('linha_base_histograma', []))
    for eq in data.get('linha_base_equipamentos', []):
        hist_all.append({'funcao': eq.get('equipamento', ''), 'tipo': 'equipamento',
                         'semanas': eq.get('semanas', {})})
    for ri, r in enumerate(hist_all, 2):
        ws3.cell(row=ri, column=1, value=r.get('funcao', '')).font  = LBL_FONT
        ws3.cell(row=ri, column=2, value=r.get('tipo',   'direto')).font = DAT_FONT
        for ci, wk in enumerate(weeks, 3):
            v = r.get('semanas', {}).get(wk)
            if v is not None:
                ws3.cell(row=ri, column=ci, value=v).font = DAT_FONT

    # ── Ações Notáveis ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Acoes')
    _head(ws4, ['Ação', 'Unidade'] + weeks)
    _widths(ws4, {'A': 32, 'B': 9}, len(weeks), pw=9)
    for ri, ac in enumerate(data.get('linha_base_acoes', []), 2):
        ws4.cell(row=ri, column=1, value=ac.get('acao',    '')).font = LBL_FONT
        ws4.cell(row=ri, column=2, value=ac.get('unidade', '')).font = DAT_FONT
        for ci, wk in enumerate(weeks, 3):
            v = ac.get('semanas', {}).get(wk)
            if v is not None:
                ws4.cell(row=ri, column=ci, value=v).font = DAT_FONT

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    ct = data.get('contratada', '').replace(' ', '_')
    cn = data.get('contrato',   '').replace(' ', '_')
    return send_file(buf, as_attachment=True,
                     download_name=f'modelo_{ct}_{cn}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Upload: importa Excel e retorna JSON com dados extraídos ──────────────────

@app.route('/admin/contrato/<path:key>/import', methods=['POST'])
def admin_contrato_import_excel(key):
    if not _admin_required():
        return jsonify({'error': 'Não autorizado'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Apenas arquivos .xlsx são aceitos'}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Erro ao abrir arquivo: {e}'}), 400

    def _fval(v):
        if v is None or str(v).strip() == '': return None
        try: return float(str(v).replace(',', '.').replace(' ', ''))
        except: return None

    def _parse_series(ws_name):
        """Lê planilha de série (Financeira/Fisica): col A = rótulo, resto = período→valor."""
        if ws_name not in wb.sheetnames: return {}
        rows   = list(wb[ws_name].iter_rows(values_only=True))
        if not rows: return {}
        header  = [str(h).strip() if h is not None else '' for h in rows[0]]
        periods = header[1:]
        parsed  = {}
        for row in rows[1:]:
            if not row or row[0] is None: continue
            label = str(row[0]).strip().lower()
            vals  = {}
            for i, p in enumerate(periods):
                if not p: continue
                v = _fval(row[i + 1] if i + 1 < len(row) else None)
                if v is not None:
                    vals[p] = v
            if label: parsed[label] = vals
        return parsed

    result = {
        'financeira': {'planejado': {}, 'forecast': {}, 'replanejados': [], 'repl_n': 0},
        'fisica':     {'planejado': {}, 'forecast': {}, 'replanejados': [], 'repl_n': 0},
        'histograma': [], 'acoes': [], 'errors': [],
    }
    try:
        for label, vals in _parse_series('Financeira').items():
            if 'planejado' in label:     result['financeira']['planejado'] = vals
            elif 'forecast' in label:    result['financeira']['forecast']  = vals
            elif 'replanejado' in label: result['financeira']['replanejados'].append(vals)
        result['financeira']['repl_n'] = min(3, len(result['financeira']['replanejados']))

        for label, vals in _parse_series('Fisica').items():
            if 'planejado' in label:     result['fisica']['planejado'] = vals
            elif 'forecast' in label:    result['fisica']['forecast']  = vals
            elif 'replanejado' in label: result['fisica']['replanejados'].append(vals)
        result['fisica']['repl_n'] = min(3, len(result['fisica']['replanejados']))

        if 'Histograma' in wb.sheetnames:
            rows = list(wb['Histograma'].iter_rows(values_only=True))
            if rows:
                hdr     = [str(h).strip() if h else '' for h in rows[0]]
                wk_keys = hdr[2:]
                for row in rows[1:]:
                    if not row or row[0] is None: continue
                    funcao = str(row[0]).strip()
                    cat    = str(row[1]).strip().lower() if len(row) > 1 and row[1] else 'direto'
                    if cat not in ('direto', 'indireto', 'equipamento'): cat = 'direto'
                    semanas = {}
                    for i, wk in enumerate(wk_keys):
                        if not wk: continue
                        v = _fval(row[i + 2] if i + 2 < len(row) else None)
                        if v is not None:
                            semanas[wk] = int(round(v))
                    if funcao:
                        result['histograma'].append({'funcao': funcao, 'tipo': cat, 'semanas': semanas})

        if 'Acoes' in wb.sheetnames:
            rows = list(wb['Acoes'].iter_rows(values_only=True))
            if rows:
                hdr     = [str(h).strip() if h else '' for h in rows[0]]
                wk_keys = hdr[2:]
                for row in rows[1:]:
                    if not row or row[0] is None: continue
                    acao    = str(row[0]).strip()
                    unidade = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    semanas = {}
                    for i, wk in enumerate(wk_keys):
                        if not wk: continue
                        v = _fval(row[i + 2] if i + 2 < len(row) else None)
                        if v is not None:
                            semanas[wk] = v
                    if acao:
                        result['acoes'].append({'acao': acao, 'unidade': unidade,
                                                'semanas': semanas, 'forecast': {}, 'replanejados': []})
    except Exception as _exc:
        app.logger.exception('Erro ao processar import Excel')
        return jsonify({'error': f'Erro ao processar arquivo: {_exc}'}), 500

    return jsonify(result)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
